# customer_invoice.py
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import sys
import subprocess
import shutil
from datetime import datetime
import re
import traceback


class CustomerInvoiceTab:
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.create_widgets()

    # ─────────────── Helper Methods ───────────────
    def _parse_display_item(self, display_str):
        """Extract English name, Urdu name, and size from display string like 'اردو (English) (Large)'"""
        size = 'Normal'
        raw_item = str(display_str).strip()
        m_size = re.search(r"\((Small|Normal|Large)\)\s*$", raw_item, flags=re.IGNORECASE)
        if m_size:
            size = m_size.group(1).capitalize()
            raw_item = re.sub(r"\s*\(%s\)\s*$" % re.escape(m_size.group(1)), '', raw_item, flags=re.IGNORECASE).strip()
        english_name = raw_item
        urdu_name = ''
        m2 = re.search(r"\(([^)]+)\)\s*$", raw_item)
        if m2:
            english_name = m2.group(1).strip()
            urdu_name = raw_item[:m2.start()].strip()
        else:
            english_name = raw_item.strip()
        return english_name, urdu_name, size

    def _find_urdu_for_english(self, english_name):
        """Look up Urdu name from app.vegetables list"""
        vegetables = getattr(self.app, 'vegetables', [])
        for v in vegetables:
            if v.get('english', '').strip().lower() == english_name.lower():
                return v.get('urdu', '')
        return ''

    def _match_rate_for_item(self, item_name):
        """Robustly match rate using English name only"""
        if not hasattr(self.app, 'rate_list') or not self.app.rate_list:
            return None
        english_name, _, _ = self._parse_display_item(item_name)
        for rate_item, rate in self.app.rate_list.items():
            if rate_item.strip().lower() == english_name.lower():
                return rate
        english_name_lower = english_name.lower().strip()
        for rate_item, rate in self.app.rate_list.items():
            rate_item_lower = rate_item.lower().strip()
            if english_name_lower in rate_item_lower or rate_item_lower in english_name_lower:
                return rate
        return None

    # ─────────────── NEW: WEB INVOICE IMPORT LOGIC ───────────────
    def import_web_invoices(self):
        """Import one or more web-generated invoices (with PKT timestamp-based numbering)."""
        invoices_dir = self.app.invoices_dir
        os.makedirs(invoices_dir, exist_ok=True)
        filenames = filedialog.askopenfilenames(
            title="Select Web Invoice(s) to Import",
            initialdir=invoices_dir,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filenames:
            return
        imported_count = 0
        errors = []
        for filename in filenames:
            try:
                self._import_single_web_invoice(filename)
                imported_count += 1
            except Exception as e:
                errors.append(f"{os.path.basename(filename)}: {str(e)}")
                traceback.print_exc()
        msg = f"✅ Successfully imported {imported_count} web invoice(s)."
        if errors:
            msg += f"\n⚠️ Errors ({len(errors)}):\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                msg += f"\n... and {len(errors) - 5} more."
            messagebox.showwarning("Import Complete", msg)
        else:
            messagebox.showinfo("Success", msg)

    def _import_single_web_invoice(self, filename):
        """Parse and integrate a single web-generated invoice."""
        invoice_date = self.app.selected_date  # fallback date
        pk_tz = None
        try:
            from pytz import timezone
            pk_tz = timezone('Asia/Karachi')
        except ImportError:
            pass  # Fallback to naive datetime if pytz not installed

        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        ws = wb.active

        # --- Extract invoice number from B2 ---
        invoice_num_from_file = None
        invoice_datetime = None
        b2_value = ws['B2'].value
        if b2_value and "Invoice #" in str(b2_value):
            try:
                num_str = str(b2_value).replace("Invoice #", "").strip()
                if num_str.isdigit() and len(num_str) == 14:
                    invoice_num_from_file = int(num_str)
                    dt_from_num = datetime.strptime(num_str, "%Y%m%d%H%M%S")
                    if pk_tz:
                        invoice_datetime = pk_tz.localize(dt_from_num)
                    else:
                        invoice_datetime = dt_from_num
            except Exception:
                pass

        # --- Extract time from E2 (e.g., "09-Jan-2026 02:30 PM") ---
        e2_value = ws['E2'].value
        if e2_value:
            try:
                parsed_time = datetime.strptime(str(e2_value), "%d-%b-%Y %I:%M %p")
                if pk_tz:
                    invoice_datetime = pk_tz.localize(parsed_time)
                else:
                    invoice_datetime = parsed_time
                invoice_date = parsed_time.strftime("%Y-%m-%d")
            except Exception:
                pass

        if invoice_datetime:
            invoice_date = invoice_datetime.strftime("%Y-%m-%d")

        # --- Extract customer info ---
        customer_name = ""
        customer_phone = ""
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if row and len(row) >= 5:
                if row[1] and "Customer Name:" in str(row[1]):
                    customer_name = str(row[2]) if row[2] else ""
                    customer_phone = str(row[4]) if row[4] else ""
                    break
            elif row[0] and "Customer Name:" in str(row[0]):
                customer_name = str(row[1]) if row[1] else ""
                customer_phone = str(row[4]) if len(row) > 4 and row[4] else ""
                break

        # --- Locate items table ---
        items_start_row = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and row[0] and str(row[0]).strip() == 'No.':
                if len(row) >= 5 and 'Item Name' in str(row[1]) and 'Quantity' in str(row[2]):
                    items_start_row = idx + 1
                    break
        if items_start_row is None:
            raise ValueError("Items table not found")

        # --- Parse items ---
        invoice_items = []
        row_idx = items_start_row
        while row_idx <= ws.max_row:
            row_cells = ws[row_idx]
            if not row_cells[0].value or "Total Amount:" in str(row_cells[0].value):
                break
            try:
                item_name = row_cells[1].value or ""
                qty_str = str(row_cells[2].value) if row_cells[2].value else "0 kg"
                rate = float(row_cells[3].value) if row_cells[3].value not in (None, "") else 0.0
                total = float(row_cells[4].value) if row_cells[4].value not in (None, "") else 0.0
                english_name, urdu_name, size = self._parse_display_item(item_name)
                if not urdu_name:
                    urdu_name = self._find_urdu_for_english(english_name)
                invoice_items.append({
                    'vegetable': english_name,
                    'urdu': urdu_name,
                    'size': size,
                    'quantity': qty_str,
                    'rate': rate,
                    'total': total
                })
            except Exception as e:
                print(f"Skipping invalid row {row_idx}: {e}")
            row_idx += 1

        if not invoice_items:
            raise ValueError("No items found")

        # --- Avoid duplicates ---
        if invoice_num_from_file:
            for inv in self.app.invoices:
                if inv.get('invoice_number') == invoice_num_from_file:
                    raise ValueError("Invoice already exists")

        # --- Create new invoice record ---
        total_amount = sum(item['total'] for item in invoice_items)
        time_display = invoice_datetime.strftime("%d-%b-%Y %I:%M %p") if invoice_datetime else datetime.now().strftime("%d-%b-%Y %I:%M %p")
        new_invoice = {
            'invoice_number': invoice_num_from_file if invoice_num_from_file else self.app.invoice_counter + 1,
            'customer_name': customer_name,
            'customer_phone': customer_phone,
            'items': invoice_items,
            'total_amount': total_amount,
            'date': invoice_date,
            'time': time_display,
            'status': 'active',
            'filepath': filename
        }

        # Only increment counter if we didn't use a timestamp ID (for legacy support)
        if not invoice_num_from_file and invoice_num_from_file != self.app.invoice_counter + 1:
            self.app.invoice_counter += 1
            self.app.save_invoice_counter()

        self.app.invoices.append(new_invoice)

        # --- Add to sales ledger ---
        for item in invoice_items:
            english_name = item['vegetable']
            urdu_name = item.get('urdu', '')
            size = item.get('size', 'Normal')
            if urdu_name and size.lower() != 'normal':
                urdu_size_map = {'Small': 'چھوٹا سائز', 'Normal': 'درمیانہ سائز', 'Large': 'بڑا سائز'}
                english_size_map = {'Small': 'small size', 'Large': 'big size'}
                urdu_display = f"{urdu_name} {urdu_size_map.get(size, '')}"
                english_display = f"{english_name} {english_size_map.get(size, size)}"
                vegetable_display = f"{urdu_display} ({english_display})"
            elif urdu_name:
                vegetable_display = f"{urdu_name} ({english_name})"
            else:
                vegetable_display = english_name

            sale = {
                'source': f"Invoice #{new_invoice['invoice_number']}",
                'vegetable': vegetable_display,
                'vegetable_english': english_name,
                'vegetable_urdu': urdu_name,
                'quantity': item['quantity'],
                'rate': item['rate'],
                'total': item['total'],
                'invoice_number': new_invoice['invoice_number']
            }
            self.app.sales.append(sale)
            if hasattr(self.app, 'sales_tree'):
                self.app.sales_tree.insert('', 'end', values=(
                    sale['source'], sale['vegetable'], sale['quantity'],
                    f"{sale['rate']:.2f}", f"{sale['total']:.2f}"
                ))

        # Update invoices tree
        if hasattr(self.app, 'invoices_tree'):
            self.app.invoices_tree.insert('', 0, values=(
                f"#{new_invoice['invoice_number']}",
                new_invoice['customer_name']
            ))

        self.app.save_sales()
        self.app.save_invoices()
        self.app.update_summary()

    # ─────────────── Core Functional Methods ───────────────
    def upload_rate_list(self):
        filename = filedialog.askopenfilename(
            title="Select Rate List Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            try:
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
                self.app.rate_list = {}
                header_found = False
                for row in ws.iter_rows(min_row=1):
                    if not header_found:
                        if row[0].value and 'Item' in str(row[0].value):
                            header_found = True
                            continue
                    if row[0].value and row[1].value:
                        item_name = str(row[0].value).strip()
                        try:
                            rate = float(row[1].value)
                            self.app.rate_list[item_name] = rate
                        except Exception:
                            continue
                self.app.rate_status_label.configure(
                    text=f"✓ Rate list loaded: {len(self.app.rate_list)} items",
                    text_color=self.app.colors['primary']
                )
                messagebox.showinfo("Success", f"Rate list loaded with {len(self.app.rate_list)} items!")
            except Exception as e:
                messagebox.showerror("Error", f"Error loading rate list: {str(e)}")

    def load_item_rate(self, event=None):
        item = self.app.invoice_item_var.get()
        if item:
            rate = self._match_rate_for_item(item)
            if rate is not None:
                self.app.invoice_rate_var.set(f"{rate:.2f}")
            else:
                english_name, urdu_name, size = self._parse_display_item(item)
                print(f"DEBUG: Could not find rate for '{item}' (parsed as '{english_name}')")
                print(f"DEBUG: Available rate list keys: {list(self.app.rate_list.keys()) if hasattr(self.app, 'rate_list') else 'No rate list'}")

    def calculate_invoice_item_total(self, *args):
        try:
            qty = float(self.app.invoice_qty_var.get()) if self.app.invoice_qty_var.get() else 0
            rate = float(self.app.invoice_rate_var.get()) if self.app.invoice_rate_var.get() else 0
            self.app.invoice_item_total_var.set(f"{qty * rate:.2f}")
        except Exception:
            self.app.invoice_item_total_var.set("0.00")

    def add_invoice_item(self):
        if not self.app.invoice_item_var.get() or not self.app.invoice_qty_var.get() or not self.app.invoice_rate_var.get():
            messagebox.showwarning("Missing Data", "Please fill all fields")
            return
        try:
            qty = float(self.app.invoice_qty_var.get())
            rate = float(self.app.invoice_rate_var.get())
            total = qty * rate
            size = getattr(self.app, 'invoice_size_var', tk.StringVar(value='Normal')).get()
            item_display = f"{self.app.invoice_item_var.get()} ({size})"
            self.app.invoice_items_tree.insert('', 'end', values=(
                item_display,
                f"{qty} {self.app.invoice_unit_var.get()}",
                f"{rate:.2f}",
                f"{total:.2f}"
            ))
            self.app.invoice_item_var.set('')
            self.app.invoice_qty_var.set('')
            self.app.invoice_rate_var.set('')
            self.app.invoice_item_total_var.set('0.00')
            self.update_invoice_total()
        except ValueError:
            messagebox.showerror("Invalid Data", "Please enter valid numbers")

    def delete_invoice_item(self):
        selection = self.app.invoice_items_tree.selection()
        if selection:
            self.app.invoice_items_tree.delete(selection)
            self.update_invoice_total()

    def clear_invoice(self):
        if messagebox.askyesno("Confirm", "Clear all items?"):
            for item in self.app.invoice_items_tree.get_children():
                self.app.invoice_items_tree.delete(item)
            self.update_invoice_total()

    def update_invoice_total(self):
        total = 0
        for item in self.app.invoice_items_tree.get_children():
            try:
                total += float(self.app.invoice_items_tree.item(item)['values'][3])
            except (ValueError, IndexError, TypeError):
                continue
        self.app.invoice_total_var.set(f"PKR {total:.2f}")

    def generate_invoice(self):
        if not self.app.invoice_items_tree.get_children():
            messagebox.showwarning("No Items", "Please add items to invoice first")
            return
        if not self.app.customer_name_var.get():
            messagebox.showwarning("Missing Info", "Please enter customer name")
            return

        invoice_date = self.app.selected_date
        try:
            invoices_dir = self.app.invoices_dir
            os.makedirs(invoices_dir, exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "FRUZY"
            ws['A1'].font = Font(name='Calibri', size=72, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color=self.app.colors['dark'].replace('#', ''),
                                      end_color=self.app.colors['dark'].replace('#', ''),
                                      fill_type="solid")
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f"{col}1"].fill = header_fill
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # 🔥 Generate PKT-based timestamp invoice number for NEW invoices
            if self.app.editing_invoice_number is not None:
                invoice_num = self.app.editing_invoice_number
            else:
                try:
                    from pytz import timezone
                    pk_tz = timezone('Asia/Karachi')
                    now_pk = datetime.now(pk_tz)
                except ImportError:
                    now_pk = datetime.now()
                invoice_num = int(now_pk.strftime("%Y%m%d%H%M%S"))

            ws['B2'] = f"Invoice #{invoice_num}"
            ws['B2'].font = Font(name='Calibri', size=11)
            ws['E2'] = now_pk.strftime("%d-%b-%Y %I:%M %p") if 'now_pk' in locals() else datetime.now().strftime("%d-%b-%Y %I:%M %p")
            ws['E2'].font = Font(name='Calibri', size=11)
            ws['E2'].alignment = Alignment(horizontal='right')

            ws['B4'] = "Customer Name:"
            ws['B4'].font = Font(name='Calibri', size=11)
            ws['C4'] = self.app.customer_name_var.get()
            ws['C4'].font = Font(name='Calibri', size=11, bold=True)
            ws['E4'] = self.app.customer_phone_var.get()
            ws['E4'].font = Font(name='Calibri', size=11, bold=True)

            header_row = 6
            headers = ['No.', 'Item Name', 'Quantity', 'Rate (PKR)', 'Total (PKR)']
            header_fill2 = PatternFill(start_color=self.app.colors['dark'].replace('#', ''),
                                       end_color=self.app.colors['dark'].replace('#', ''),
                                       fill_type="solid")
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.fill = header_fill2
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            row = header_row + 1
            total = 0.0
            idx = 1
            for item in self.app.invoice_items_tree.get_children():
                values = self.app.invoice_items_tree.item(item)['values']
                raw_item = str(values[0])
                english_name, urdu_name, size = self._parse_display_item(raw_item)
                if not urdu_name:
                    urdu_name = self._find_urdu_for_english(english_name)

                urdu_size_map = {'Small': 'چھوٹا سائز', 'Normal': 'درمیانہ سائز', 'Large': 'بڑا سائز'}
                english_size_map = {'Small': 'small size', 'Large': 'big size'}
                english_plain = english_name
                if size.lower() != 'normal':
                    english_plain = f"{english_name} {english_size_map.get(size, size)}"
                urdu_cell = urdu_name
                if urdu_name and size.lower() != 'normal':
                    urdu_cell = f"{urdu_name} {urdu_size_map.get(size, '')}"

                if urdu_name:
                    item_cell_value = f"{urdu_cell} ({english_plain})"
                else:
                    item_cell_value = english_plain

                ws.cell(row=row, column=1, value=idx).border = border
                cell = ws.cell(row=row, column=2, value=item_cell_value)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = border
                ws.cell(row=row, column=3, value=values[1]).border = border

                try:
                    rate_cell = ws.cell(row=row, column=4, value=float(values[2]))
                    rate_cell.number_format = '#,##0.00'
                    rate_cell.border = border
                except Exception:
                    rc = ws.cell(row=row, column=4, value=values[2])
                    rc.border = border

                try:
                    tot_cell = ws.cell(row=row, column=5, value=float(values[3]))
                    tot_cell.number_format = '#,##0.00'
                    tot_cell.border = border
                    total += float(values[3])
                except Exception:
                    tc = ws.cell(row=row, column=5, value=values[3])
                    tc.border = border

                row += 1
                idx += 1

            ws.cell(row=row + 1, column=1, value="Total Amount:").font = Font(name='Calibri', size=11, bold=True)
            total_cell = ws.cell(row=row + 1, column=5, value=total)
            total_cell.font = Font(name='Calibri', size=11, bold=True)
            total_cell.number_format = '#,##0.00'

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            # 🔥 Save NEW FILE FIRST (CRITICAL FIX)
            customer_name_clean = ''.join(c for c in self.app.customer_name_var.get() if c.isalnum()) or "Customer"
            filename = f"Invoice_{invoice_num}_{customer_name_clean}.xlsx"
            new_file_path = os.path.join(invoices_dir, filename)
            wb.save(new_file_path)  # ✅ Save before deleting old

            # Build invoice items list
            invoice_items = []
            for item in self.app.invoice_items_tree.get_children():
                values = self.app.invoice_items_tree.item(item)['values']
                english_name, urdu_name, size = self._parse_display_item(values[0])
                if not urdu_name:
                    urdu_name = self._find_urdu_for_english(english_name)
                invoice_items.append({
                    'vegetable': english_name,
                    'urdu': urdu_name,
                    'size': size,
                    'quantity': values[1],
                    'rate': float(values[2]),
                    'total': float(values[3])
                })

            # Create new invoice record with NEW filepath
            new_invoice = {
                'invoice_number': invoice_num,
                'customer_name': self.app.customer_name_var.get(),
                'customer_phone': self.app.customer_phone_var.get(),
                'items': invoice_items,
                'total_amount': total,
                'date': invoice_date,
                'time': now_pk.strftime("%d-%b-%Y %I:%M %p") if 'now_pk' in locals() else datetime.now().strftime("%d-%b-%Y %I:%M %p"),
                'status': 'active',
                'filepath': new_file_path  # ✅ Points to new file
            }

            # Handle editing vs new
            if self.app.editing_invoice_number is not None:
                # Remove old sales entries
                self.app.sales = [s for s in self.app.sales if s.get('invoice_number') != self.app.editing_invoice_number]

                # Delete OLD file ONLY AFTER new is saved
                old_file_path = None
                for inv in self.app.invoices:
                    if inv.get('invoice_number') == self.app.editing_invoice_number:
                        old_file_path = inv.get('filepath')
                        break

                if old_file_path and os.path.exists(old_file_path):
                    # Avoid self-delete if same path (shouldn't happen, but safe)
                    if os.path.abspath(old_file_path) != os.path.abspath(new_file_path):
                        try:
                            os.remove(old_file_path)
                        except Exception as e:
                            print(f"Warning: Could not delete old invoice file: {e}")

                # Replace in memory
                for i, inv in enumerate(self.app.invoices):
                    if inv.get('invoice_number') == self.app.editing_invoice_number:
                        self.app.invoices[i] = new_invoice
                        break

                # Clean UI sales tree
                if hasattr(self.app, 'sales_tree'):
                    items_to_delete = []
                    for item_id in self.app.sales_tree.get_children():
                        values = self.app.sales_tree.item(item_id)['values']
                        source = values[0] if values else ""
                        if f"Invoice #{self.app.editing_invoice_number}" == source:
                            items_to_delete.append(item_id)
                    for item_id in items_to_delete:
                        self.app.sales_tree.delete(item_id)

            else:
                # New invoice
                self.app.invoices.append(new_invoice)

            # Update UI trees
            if hasattr(self.app, 'invoices_tree'):
                if self.app.editing_invoice_number is not None:
                    for item_id in self.app.invoices_tree.get_children():
                        vals = self.app.invoices_tree.item(item_id)['values']
                        if vals and f"#{self.app.editing_invoice_number}" == vals[0]:
                            self.app.invoices_tree.item(item_id, values=(
                                f"#{invoice_num}",
                                new_invoice['customer_name']
                            ))
                            break
                else:
                    self.app.invoices_tree.insert('', 0, values=(
                        f"#{invoice_num}",
                        self.app.customer_name_var.get()
                    ))

            # Add to sales ledger
            for item in invoice_items:
                english_name = item['vegetable']
                urdu_name = item.get('urdu', '')
                size = item.get('size', 'Normal')
                if urdu_name and size.lower() != 'normal':
                    urdu_size_map = {'Small': 'چھوٹا سائز', 'Normal': 'درمیانہ سائز', 'Large': 'بڑا سائز'}
                    english_size_map = {'Small': 'small size', 'Large': 'big size'}
                    urdu_display = f"{urdu_name} {urdu_size_map.get(size, '')}"
                    english_display = f"{english_name} {english_size_map.get(size, size)}"
                    vegetable_display = f"{urdu_display} ({english_display})"
                elif urdu_name:
                    vegetable_display = f"{urdu_name} ({english_name})"
                else:
                    vegetable_display = english_name

                sale = {
                    'source': f'Invoice #{invoice_num}',
                    'vegetable': vegetable_display,
                    'vegetable_english': english_name,
                    'vegetable_urdu': urdu_name,
                    'quantity': item['quantity'],
                    'rate': f"{item['rate']:.2f}",
                    'total': f"{item['total']:.2f}",
                    'invoice_number': invoice_num
                }
                self.app.sales.append(sale)
                if hasattr(self.app, 'sales_tree'):
                    self.app.sales_tree.insert('', 'end', values=(
                        sale['source'], sale['vegetable'], sale['quantity'],
                        sale['rate'], sale['total']
                    ))

            # Reset form
            self.app.editing_invoice_number = None
            self.app.customer_name_var.set('')
            self.app.customer_phone_var.set('')
            self.clear_invoice()
            self.app.save_sales()
            self.app.save_invoices()
            if hasattr(self.app, 'sales_tab_instance') and self.app.sales_tab_instance:
                try:
                    self.app.sales_tab_instance.reload_sales_list()
                except Exception as e:
                    print(f"Error refreshing sales list: {e}")
            self.app.update_summary()

            date_obj = datetime.strptime(invoice_date, "%Y-%m-%d")
            date_display = date_obj.strftime("%A, %B %d, %Y")
            messagebox.showinfo("Success", f"Invoice #{invoice_num} generated for {date_display}\nSaved as:\n{new_file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate invoice: {str(e)}")
            traceback.print_exc()

    def open_invoice_folder(self):
        invoices_dir = self.app.invoices_dir
        try:
            os.makedirs(invoices_dir, exist_ok=True)
            if sys.platform == "darwin":
                subprocess.run(["open", invoices_dir], check=False)
            elif os.name == "nt":
                os.startfile(invoices_dir)
            else:
                if shutil.which("xdg-open"):
                    subprocess.run(["xdg-open", invoices_dir], check=False)
                else:
                    messagebox.showinfo("Invoice Folder", f"Invoice folder:\n{invoices_dir}")
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open folder: {e}")

    def open_invoice(self, invoice_number=None):
        if invoice_number is None:
            selection = self.app.invoices_tree.selection()
            if not selection:
                return
            selected_item = selection[0]
            item_values = self.app.invoices_tree.item(selected_item)['values']
            if not item_values:
                return
            inv_num_str = item_values[0].replace('#', '')
            try:
                invoice_number = int(inv_num_str)
            except ValueError:
                return

        target_invoice = None
        for inv in self.app.invoices:
            if inv.get('invoice_number') == invoice_number:
                target_invoice = inv
                break

        if not target_invoice or 'filepath' not in target_invoice:
            messagebox.showerror("Error", "Invoice file not found.")
            return

        filepath = target_invoice['filepath']
        if not os.path.exists(filepath):
            messagebox.showerror("Error", "Invoice file has been moved or deleted.")
            return

        try:
            if sys.platform == "darwin":
                subprocess.run(["open", filepath], check=False)
            elif os.name == "nt":
                os.startfile(filepath)
            else:
                if shutil.which("xdg-open"):
                    subprocess.run(["xdg-open", filepath], check=False)
                else:
                    messagebox.showinfo("Invoice", f"File:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open invoice: {str(e)}")

    def reimport_invoice(self):
        invoices_dir = self.app.invoices_dir
        os.makedirs(invoices_dir, exist_ok=True)
        filename = filedialog.askopenfilename(
            title="Select Invoice to Re-import",
            initialdir=invoices_dir,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filename:
            return

        invoice_date = self.app.selected_date
        try:
            wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
            ws = wb.active

            customer_name = ""
            customer_phone = ""
            invoice_num_from_file = None

            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                if row[0] and "Invoice #" in str(row[0]):
                    try:
                        invoice_num_from_file = int(str(row[0]).replace("Invoice #", "").strip())
                    except (ValueError, IndexError):
                        pass
                    break

            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if not row or len(row) < 2:
                    continue
                if (len(row) > 1 and row[1] and "Customer Name:" in str(row[1])) or (row[0] and "Customer Name:" in str(row[0])):
                    if len(row) > 1 and row[1] and "Customer Name:" in str(row[1]):
                        customer_name = str(row[2]) if len(row) > 2 and row[2] else ""
                        customer_phone = str(row[4]) if len(row) > 4 and row[4] else ""
                    else:
                        customer_name = str(row[1]) if len(row) > 1 and row[1] else ""
                        customer_phone = str(row[4]) if len(row) > 4 and row[4] else ""
                    break

            items_start_row = None
            headers = ['No.', 'Item Name', 'Quantity', 'Rate (PKR)', 'Total (PKR)']
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row[0] and str(row[0]).strip() == 'No.':
                    if all(h in str(cell) for h, cell in zip(headers, row[:5])):
                        items_start_row = idx + 1
                        break
            if items_start_row is None:
                messagebox.showerror("Error", "Could not locate invoice items in the file.")
                return

            invoice_items = []
            row_idx = items_start_row
            while row_idx <= ws.max_row:
                row = ws[row_idx]
                if not row[0].value or "Total Amount:" in str(row[0].value):
                    break
                try:
                    item_name = row[1].value or ""
                    qty = row[2].value or ""
                    rate = float(row[3].value) if row[3].value not in (None, "") else 0.0
                    total = float(row[4].value) if row[4].value not in (None, "") else 0.0
                    english_name, urdu_name, size = self._parse_display_item(item_name)
                    if not urdu_name:
                        urdu_name = self._find_urdu_for_english(english_name)
                    invoice_items.append({
                        'vegetable': english_name,
                        'urdu': urdu_name,
                        'size': size,
                        'quantity': str(qty),
                        'rate': rate,
                        'total': total
                    })
                except Exception:
                    pass
                row_idx += 1

            if invoice_num_from_file is None:
                basename = os.path.basename(filename)
                try:
                    invoice_num_from_file = int(basename.split('_')[1])
                except (IndexError, ValueError):
                    pass

            existing_invoice = None
            if invoice_num_from_file is not None:
                for inv in self.app.invoices:
                    if inv.get('invoice_number') == invoice_num_from_file:
                        existing_invoice = inv
                        break

            if existing_invoice:
                self.app.customer_name_var.set(existing_invoice.get('customer_name', ''))
                self.app.customer_phone_var.set(existing_invoice.get('customer_phone', ''))
                for item in self.app.invoice_items_tree.get_children():
                    self.app.invoice_items_tree.delete(item)
                for item in existing_invoice.get('items', []):
                    item_name = item.get('vegetable', '')
                    size = item.get('size', 'Normal')
                    quantity = item.get('quantity', '')
                    rate = item.get('rate', 0)
                    display_name = f"{item_name} ({size})"
                    self.app.invoice_items_tree.insert('', 'end', values=(
                        display_name,
                        quantity,
                        f"{rate:.2f}",
                        f"{item.get('total', 0):.2f}"
                    ))
                self.update_invoice_total()
                self.app.editing_invoice_number = invoice_num_from_file
                messagebox.showinfo("Invoice Loaded", f"Invoice #{invoice_num_from_file} loaded for editing.\nCustomer: {existing_invoice.get('customer_name', '')}")
            else:
                total_amount = sum(item['total'] for item in invoice_items)
                new_invoice = {
                    'invoice_number': invoice_num_from_file if invoice_num_from_file else self.app.invoice_counter + 1,
                    'customer_name': customer_name,
                    'customer_phone': customer_phone,
                    'items': invoice_items,
                    'total_amount': total_amount,
                    'date': invoice_date,
                    'time': datetime.now().strftime("%d-%b-%Y %I:%M %p"),
                    'status': 'active',
                    'filepath': filename
                }
                if invoice_num_from_file is None:
                    new_invoice['invoice_number'] = self.app.invoice_counter + 1
                    self.app.invoice_counter += 1
                    self.app.save_invoice_counter()
                self.app.invoices.append(new_invoice)

                if hasattr(self.app, 'invoices_tree'):
                    self.app.invoices_tree.insert('', 0, values=(
                        f"#{new_invoice['invoice_number']}",
                        new_invoice['customer_name']
                    ))

                for item in invoice_items:
                    english_name = item['vegetable']
                    urdu_name = item.get('urdu', '')
                    size = item.get('size', 'Normal')
                    if urdu_name and size.lower() != 'normal':
                        urdu_size_map = {'Small': 'چھوٹا سائز', 'Normal': 'درمیانہ سائز', 'Large': 'بڑا سائز'}
                        english_size_map = {'Small': 'small size', 'Large': 'big size'}
                        urdu_display = f"{urdu_name} {urdu_size_map.get(size, '')}"
                        english_display = f"{english_name} ({english_size_map.get(size, size)})"
                        vegetable_display = f"{urdu_display} ({english_display})"
                    elif urdu_name:
                        vegetable_display = f"{urdu_name} ({english_name})"
                    else:
                        vegetable_display = english_name

                    sale = {
                        'source': f"Invoice #{new_invoice['invoice_number']}",
                        'vegetable': vegetable_display,
                        'vegetable_english': english_name,
                        'vegetable_urdu': urdu_name,
                        'quantity': item['quantity'],
                        'rate': item['rate'],
                        'total': item['total'],
                        'invoice_number': new_invoice['invoice_number']
                    }
                    self.app.sales.append(sale)
                    if hasattr(self.app, 'sales_tree'):
                        self.app.sales_tree.insert('', 'end', values=(
                            sale['source'], sale['vegetable'], sale['quantity'],
                            f"{sale['rate']:.2f}", f"{sale['total']:.2f}"
                        ))

                self.app.save_sales()
                self.app.save_invoices()
                self.app.update_summary()

                self.app.customer_name_var.set(customer_name)
                self.app.customer_phone_var.set(customer_phone)
                for item in self.app.invoice_items_tree.get_children():
                    self.app.invoice_items_tree.delete(item)
                for item in invoice_items:
                    self.app.invoice_items_tree.insert('', 'end', values=(
                        item['vegetable'],
                        item['quantity'],
                        f"{item['rate']:.2f}",
                        f"{item['total']:.2f}"
                    ))
                self.update_invoice_total()
                self.app.editing_invoice_number = new_invoice['invoice_number']
                messagebox.showinfo("Success", f"Invoice #{new_invoice['invoice_number']} re-imported successfully.\nLoaded for editing as: {customer_name}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to re-import invoice:\n{str(e)}")
            traceback.print_exc()

    def open_edit_invoice_window(self, invoice):
        self.app.editing_invoice_number = invoice.get('invoice_number')
        self.app.customer_name_var.set(invoice.get('customer_name', ''))
        self.app.customer_phone_var.set(invoice.get('customer_phone', ''))
        for item in self.app.invoice_items_tree.get_children():
            self.app.invoice_items_tree.delete(item)

        update_rates = False
        if getattr(self.app, 'rate_list', None):
            result = messagebox.askyesno(
                "Update Rates",
                f"Found {len(self.app.rate_list)} items in the current rate list.\n"
                "Would you like to automatically update all rates from the current rate list?\n"
                "(You can manually edit any rate afterwards if needed.)",
                icon='question'
            )
            update_rates = result

        for item in invoice.get('items', []):
            item_name = item.get('vegetable', '')
            size = item.get('size', 'Normal')
            quantity = item.get('quantity', '')
            rate = item.get('rate', 0)
            if update_rates:
                matched_rate = self._match_rate_for_item(item_name)
                if matched_rate is not None:
                    rate = matched_rate
            try:
                qty_num = float(quantity.split()[0]) if isinstance(quantity, str) else float(quantity)
                total = qty_num * rate
            except (ValueError, IndexError, AttributeError):
                total = item['total']
            display_name = f"{item_name} ({size})"
            self.app.invoice_items_tree.insert('', 'end', values=(
                display_name,
                quantity,
                f"{rate:.2f}",
                f"{total:.2f}"
            ))
        self.update_invoice_total()

        if update_rates:
            messagebox.showinfo(
                "Rates Updated",
                f"Invoice #{invoice.get('invoice_number')} loaded with updated rates.\n"
                "All matching items have been updated from the current rate list.\n"
                "You can edit any rates manually if needed before saving."
            )
        else:
            messagebox.showinfo(
                "Edit Mode",
                f"Invoice #{invoice.get('invoice_number')} loaded for editing.\n"
                "Original rates have been preserved.\n"
                "To use current rates, please reload the rate list first."
            )

    def edit_invoice_dialog(self):
        if not self.app.invoices:
            messagebox.showwarning("No Invoices", "No invoices available to edit")
            return
        available_invoices = [inv for inv in self.app.invoices if inv.get('status') == 'active']
        if not available_invoices:
            messagebox.showwarning("No Active Invoices", "No active invoices found")
            return

        dialog = ctk.CTkToplevel(self.app.root)
        dialog.title("Edit Invoice")
        dialog.geometry("400x300")
        dialog.resizable(False, False)
        dialog.transient(self.app.root)
        dialog.grab_set()

        ctk.CTkLabel(dialog, text="Select Invoice to Edit:",
                     font=('Arial', 14, 'bold')).pack(pady=20)

        listbox_frame = ctk.CTkFrame(dialog)
        listbox_frame.pack(fill='both', expand=True, padx=20, pady=10)

        scrollbar = ttk.Scrollbar(listbox_frame)
        scrollbar.pack(side='right', fill='y')

        invoice_listbox = tk.Listbox(listbox_frame, font=('Arial', 10),
                                     yscrollcommand=scrollbar.set,
                                     selectbackground=self.app.colors['primary'],
                                     selectforeground='white')
        invoice_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=invoice_listbox.yview)

        for invoice in available_invoices:
            inv_num = invoice.get('invoice_number', '?')
            customer = invoice.get('customer_name', 'Unknown')
            invoice_listbox.insert(tk.END, f"Invoice #{inv_num} - {customer}")

        def on_select():
            try:
                selection = invoice_listbox.curselection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select an invoice")
                    return
                selected_idx = selection[0]
                selected_invoice = available_invoices[selected_idx]
                self.open_edit_invoice_window(selected_invoice)
                dialog.destroy()
            except Exception as e:
                traceback.print_exc()
                messagebox.showerror("Error", f"An error occurred: {str(e)}")

        def on_cancel():
            dialog.destroy()

        button_frame = ctk.CTkFrame(dialog)
        button_frame.pack(fill='x', padx=20, pady=15)
        ctk.CTkButton(button_frame, text="Edit", font=('Arial', 12, 'bold'),
                      command=on_select).pack(side='left', padx=5)
        ctk.CTkButton(button_frame, text="Cancel", font=('Arial', 12, 'bold'),
                      command=on_cancel).pack(side='left', padx=5)

    def show_invoice_item_context_menu(self, event):
        tree = self.app.invoice_items_tree
        item_id = tree.identify_row(event.y)
        if not item_id:
            try:
                rel_y = int(event.y_root - tree.winfo_rooty())
                item_id = tree.identify_row(rel_y)
            except Exception:
                return
        if not item_id:
            return
        tree.selection_set(item_id)
        context_menu = tk.Menu(self.app.root, tearoff=0)
        context_menu.add_command(label="✏️ Edit Item", command=lambda: self.edit_invoice_item(event))
        context_menu.add_command(label="🗑️ Delete Item", command=self.delete_invoice_item)
        context_menu.add_separator()
        context_menu.add_command(label="🗑️ Clear All", command=self.clear_invoice)
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def edit_invoice_item(self, event):
        selection = self.app.invoice_items_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an item to edit")
            return
        selected_item = selection[0]
        values = self.app.invoice_items_tree.item(selected_item)['values']

        dialog = ctk.CTkToplevel(self.app.root)
        dialog.title("Edit Invoice Item")
        dialog.geometry("500x300")
        dialog.resizable(False, False)
        dialog.transient(self.app.root)
        dialog.grab_set()

        raw_name = str(values[0])
        size_val = 'Normal'
        m_size = re.search(r"\((Small|Normal|Large)\)\s*$", raw_name, flags=re.IGNORECASE)
        if m_size:
            size_val = m_size.group(1).capitalize()
            raw_name = re.sub(r"\s*\(%s\)\s*$" % re.escape(m_size.group(1)), '', raw_name, flags=re.IGNORECASE)

        item_var = tk.StringVar(value=raw_name)
        ctk.CTkLabel(dialog, text="Item Name:", font=('Arial', 12, 'bold')).grid(row=0, column=0, padx=15, pady=10, sticky='w')
        item_entry = ctk.CTkEntry(dialog, textvariable=item_var, font=('Arial', 12), width=300)
        item_entry.grid(row=0, column=1, padx=15, pady=10, sticky='ew')

        ctk.CTkLabel(dialog, text="Size:", font=('Arial', 12, 'bold')).grid(row=0, column=2, padx=5, pady=10, sticky='w')
        size_var = tk.StringVar(value=size_val)
        size_combo = ttk.Combobox(dialog, textvariable=size_var, values=['Small', 'Normal', 'Large'], width=12, state='readonly')
        size_combo.grid(row=0, column=3, padx=5, pady=10, sticky='w')

        qty_var = tk.StringVar(value=values[1])
        ctk.CTkLabel(dialog, text="Quantity:", font=('Arial', 12, 'bold')).grid(row=1, column=0, padx=15, pady=10, sticky='w')
        qty_entry = ctk.CTkEntry(dialog, textvariable=qty_var, font=('Arial', 12), width=300)
        qty_entry.grid(row=1, column=1, padx=15, pady=10, sticky='ew')

        rate_var = tk.StringVar(value=values[2])
        ctk.CTkLabel(dialog, text="Rate (PKR):", font=('Arial', 12, 'bold')).grid(row=2, column=0, padx=15, pady=10, sticky='w')
        rate_entry = ctk.CTkEntry(dialog, textvariable=rate_var, font=('Arial', 12), width=300)
        rate_entry.grid(row=2, column=1, padx=15, pady=10, sticky='ew')

        total_var = tk.StringVar(value=values[3])
        ctk.CTkLabel(dialog, text="Total (Auto):", font=('Arial', 12, 'bold')).grid(row=3, column=0, padx=15, pady=10, sticky='w')
        total_label = ctk.CTkLabel(dialog, textvariable=total_var, font=('Arial', 12, 'bold'))
        total_label.grid(row=3, column=1, padx=15, pady=10, sticky='ew')

        def update_total(*args):
            try:
                qty_str = qty_var.get().split()[0] if qty_var.get() else '0'
                qty = float(qty_str)
                rate = float(rate_var.get()) if rate_var.get() else 0
                total = qty * rate
                total_var.set(f"{total:.2f}")
            except ValueError:
                total_var.set("0.00")

        qty_var.trace('w', update_total)
        rate_var.trace('w', update_total)

        def save_changes():
            try:
                display_name = f"{item_var.get()} ({size_var.get()})"
                self.app.invoice_items_tree.item(selected_item, values=(
                    display_name,
                    qty_var.get(),
                    f"{float(rate_var.get()):.2f}",
                    total_var.get()
                ))
                self.update_invoice_total()
                dialog.destroy()
                messagebox.showinfo("Success", "Item updated successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update item: {str(e)}")

        def cancel_edit():
            dialog.destroy()

        button_frame = ctk.CTkFrame(dialog)
        button_frame.grid(row=4, column=0, columnspan=2, padx=15, pady=20)
        ctk.CTkButton(button_frame, text="Save Changes", font=('Arial', 12, 'bold'),
                      command=save_changes).pack(side='left', padx=5)
        ctk.CTkButton(button_frame, text="Cancel", font=('Arial', 12, 'bold'),
                      command=cancel_edit).pack(side='left', padx=5)
        dialog.grid_columnconfigure(1, weight=1)

    def update_invoice_rates_from_current_list(self):
        if not getattr(self.app, 'rate_list', None):
            messagebox.showwarning("No Rate List", "Please upload a rate list first.")
            return
        updated = False
        for item_id in self.app.invoice_items_tree.get_children():
            values = self.app.invoice_items_tree.item(item_id)['values']
            item_name = values[0]
            quantity_str = values[1]
            current_rate = float(values[2]) if values[2] else 0.0
            new_rate = self._match_rate_for_item(item_name)
            if new_rate is None:
                continue
            if abs(new_rate - current_rate) > 1e-6:
                try:
                    qty_num = float(quantity_str.split()[0]) if isinstance(quantity_str, str) else float(quantity_str)
                    new_total = qty_num * new_rate
                    self.app.invoice_items_tree.item(item_id, values=(
                        item_name,
                        quantity_str,
                        f"{new_rate:.2f}",
                        f"{new_total:.2f}"
                    ))
                    updated = True
                except (ValueError, IndexError):
                    continue
        if updated:
            self.update_invoice_total()
            messagebox.showinfo("Success", "All matching item rates updated to today's rates!")
        else:
            messagebox.showinfo("No Changes", "No items matched the current rate list, or rates are unchanged.")

    # ─────────────── Context Menu for Recent Invoices (SIMPLIFIED) ───────────────
    def show_invoice_context_menu(self, event):
        tree = self.app.invoices_tree
        item_id = tree.identify_row(event.y)
        if not item_id:
            try:
                rel_y = int(event.y_root - tree.winfo_rooty())
                item_id = tree.identify_row(rel_y)
            except Exception:
                return
        if not item_id:
            return
        tree.selection_set(item_id)
        context_menu = tk.Menu(self.app.root, tearoff=0)
        context_menu.add_command(label="📖 Open Invoice", command=self._context_open_invoice)
        context_menu.add_command(label="🗑️ Delete Invoice", command=self._context_delete_invoice)
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def _context_open_invoice(self):
        selection = self.app.invoices_tree.selection()
        if not selection:
            return
        item_values = self.app.invoices_tree.item(selection[0])['values']
        if not item_values:
            return
        try:
            inv_num = int(str(item_values[0]).replace('#', ''))
            self.open_invoice(invoice_number=inv_num)
        except (ValueError, IndexError):
            pass

    def _context_delete_invoice(self):
        selection = self.app.invoices_tree.selection()
        if not selection:
            return
        item_values = self.app.invoices_tree.item(selection[0])['values']
        if not item_values:
            return
        try:
            inv_num = int(str(item_values[0]).replace('#', ''))
            if not messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete Invoice #{inv_num}?\nThis action cannot be undone."):
                return
            for i, inv in enumerate(self.app.invoices):
                if inv.get('invoice_number') == inv_num:
                    filepath = inv.get('filepath')
                    if filepath and os.path.exists(filepath):
                        try:
                            os.remove(filepath)
                        except Exception as e:
                            messagebox.showwarning("File Error", f"Could not delete file:\n{str(e)}")
                    if self.app.editing_invoice_number == inv_num:
                        for item in self.app.invoice_items_tree.get_children():
                            self.app.invoice_items_tree.delete(item)
                        self.app.customer_name_var.set("")
                        self.app.customer_phone_var.set("")
                        self.app.invoice_total_var.set("PKR 0.00")
                        self.app.editing_invoice_number = None
                    self.app.sales = [s for s in self.app.sales if s.get('invoice_number') != inv_num]
                    if hasattr(self.app, 'sales_tree'):
                        for item_id in list(self.app.sales_tree.get_children()):
                            values = self.app.sales_tree.item(item_id)['values']
                            if values and f"Invoice #{inv_num}" in str(values[0]):
                                self.app.sales_tree.delete(item_id)
                    self.app.invoices.pop(i)
                    self.app.invoices_tree.delete(selection[0])
                    self.app.save_sales()
                    self.app.save_invoices()
                    self.app.update_summary()
                    messagebox.showinfo("Success", f"Invoice #{inv_num} has been deleted.")
                    return
            messagebox.showwarning("Not Found", "Selected invoice not found.")
        except (ValueError, IndexError):
            pass

    # ─────────────── DOUBLE-CLICK HANDLER FOR EDITING ───────────────
    def _on_invoice_double_click(self, event):
        """Load selected invoice into the form for editing on double-click."""
        tree = self.app.invoices_tree
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        item_id = tree.identify_row(event.y)
        if not item_id:
            return
        tree.selection_set(item_id)
        tree.focus(item_id)
        values = tree.item(item_id)['values']
        if not values or not values[0]:
            return
        try:
            inv_num = int(str(values[0]).replace('#', ''))
            for inv in self.app.invoices:
                if inv.get('invoice_number') == inv_num and inv.get('status') == 'active':
                    self.open_edit_invoice_window(inv)
                    return
            messagebox.showwarning("Not Found", f"Invoice #{inv_num} not found or already deleted.")
        except (ValueError, IndexError, AttributeError) as e:
            messagebox.showerror("Error", "Invalid invoice selection.")

    # ─────────────── UI Creation ───────────────
    def create_widgets(self):
        ctk.set_appearance_mode("Light")
        ctk.set_default_color_theme("blue")

        main_frame = ctk.CTkFrame(self.parent, fg_color="transparent")
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        main_frame.grid_columnconfigure(0, weight=9)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_rowconfigure(0, weight=0)
        main_frame.grid_rowconfigure(1, weight=1)

        # TOP SECTION
        rate_frame = ctk.CTkFrame(main_frame, fg_color=("#f0f0f0", "#2b2b2b"))
        rate_frame.grid(row=0, column=0, columnspan=2, sticky='ew', padx=0, pady=(0, 15))

        rate_header_frame = ctk.CTkFrame(rate_frame, fg_color="transparent")
        rate_header_frame.pack(fill='x', padx=15, pady=(12, 12))

        left_header = ctk.CTkFrame(rate_header_frame, fg_color="transparent")
        left_header.pack(side='left')
        ctk.CTkLabel(left_header, text="📋 Daily Invoice List", font=('Arial', 14, 'bold')).pack(anchor='w', padx=0, pady=0)

        right_buttons = ctk.CTkFrame(rate_header_frame, fg_color="transparent")
        right_buttons.pack(side='right')

        # 🔥 ADD THE NEW BUTTON HERE
        ctk.CTkButton(right_buttons, text="💾 Re-import Invoice", width=130,
                      command=self.reimport_invoice).pack(side='left', padx=3)
        ctk.CTkButton(right_buttons, text="🌐 Import Web Invoices", width=150,
                      command=self.import_web_invoices).pack(side='left', padx=3)
        ctk.CTkButton(right_buttons, text="📋 Import Rate List", width=140,
                      command=self.upload_rate_list).pack(side='left', padx=3)

        self.app.rate_status_label = ctk.CTkLabel(rate_frame, text="No rate list loaded",
                                                  text_color="red", font=('Arial', 10))
        self.app.rate_status_label.pack(pady=(0, 10))

        # LEFT COLUMN: Invoice Creation
        left_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        left_frame.grid(row=1, column=0, sticky='nsew', padx=(0, 5))

        invoice_frame = ctk.CTkFrame(left_frame, fg_color=("#ffffff", "#1a1a1a"))
        invoice_frame.pack(fill='both', expand=True, padx=0, pady=0)

        invoice_header_frame = ctk.CTkFrame(invoice_frame, fg_color="transparent")
        invoice_header_frame.pack(fill='x', padx=15, pady=(15, 12))
        ctk.CTkLabel(invoice_header_frame, text="🧾 Create Customer Invoice", font=('Arial', 14, 'bold')).pack(anchor='w', padx=0, pady=0)

        customer_row = ctk.CTkFrame(invoice_frame, fg_color="transparent")
        customer_row.pack(fill='x', padx=15, pady=8)
        ctk.CTkLabel(customer_row, text="Customer Name:", width=120, anchor='w').pack(side='left', padx=5)
        self.app.customer_name_var = tk.StringVar()
        ctk.CTkEntry(customer_row, textvariable=self.app.customer_name_var, width=300).pack(side='left', padx=5, fill='x', expand=True)
        ctk.CTkLabel(customer_row, text="Phone:", width=60, anchor='w').pack(side='left', padx=(10, 5))
        self.app.customer_phone_var = tk.StringVar()
        ctk.CTkEntry(customer_row, textvariable=self.app.customer_phone_var, width=150).pack(side='left', padx=5)

        items_section = ctk.CTkFrame(invoice_frame, fg_color="transparent")
        items_section.pack(fill='both', expand=True, padx=15, pady=15)

        add_item_frame = ctk.CTkFrame(items_section, fg_color=("#f9f9f9", "#252525"))
        add_item_frame.pack(fill='x', pady=(0, 15), padx=10, ipady=10)

        ctk.CTkLabel(add_item_frame, text="Item:", width=60, anchor='w').grid(row=0, column=0, padx=8, pady=8, sticky='w')
        self.app.invoice_item_var = tk.StringVar()
        item_combo = ttk.Combobox(add_item_frame, textvariable=self.app.invoice_item_var,
                                  values=[f"{v['urdu']} ({v['english']})" for v in getattr(self.app, 'vegetables', [])],
                                  font=('Arial', 11), width=50)
        item_combo.grid(row=0, column=1, columnspan=7, padx=8, pady=8, sticky='ew')
        item_combo.bind('<<ComboboxSelected>>', self.load_item_rate)

        ctk.CTkLabel(add_item_frame, text="Qty:", width=35, anchor='w').grid(row=1, column=0, padx=4, pady=6, sticky='w')
        self.app.invoice_qty_var = tk.StringVar()
        self.app.invoice_qty_var.trace('w', self.calculate_invoice_item_total)
        qty_entry = ctk.CTkEntry(add_item_frame, textvariable=self.app.invoice_qty_var, width=70)
        qty_entry.grid(row=1, column=1, padx=2, pady=6, sticky='ew')

        ctk.CTkLabel(add_item_frame, text="Unit:", width=35, anchor='w').grid(row=1, column=2, padx=4, pady=6, sticky='w')
        self.app.invoice_unit_var = tk.StringVar(value='kg')
        unit_combo = ttk.Combobox(add_item_frame, textvariable=self.app.invoice_unit_var,
                                  values=['kg', 'piece', 'dozen', 'bundle'],
                                  font=('Arial', 10), width=10, state='readonly')
        unit_combo.grid(row=1, column=3, padx=2, pady=6, sticky='ew')

        ctk.CTkLabel(add_item_frame, text="Size:", width=35, anchor='w').grid(row=1, column=4, padx=4, pady=6, sticky='w')
        self.app.invoice_size_var = tk.StringVar(value='Normal')
        size_combo = ttk.Combobox(add_item_frame, textvariable=self.app.invoice_size_var,
                                  values=['Small', 'Normal', 'Large'],
                                  font=('Arial', 10), width=10, state='readonly')
        size_combo.grid(row=1, column=5, padx=2, pady=6, sticky='ew')

        ctk.CTkLabel(add_item_frame, text="Rate:", width=35, anchor='w').grid(row=1, column=6, padx=4, pady=6, sticky='w')
        self.app.invoice_rate_var = tk.StringVar()
        self.app.invoice_rate_var.trace('w', self.calculate_invoice_item_total)
        rate_entry = ctk.CTkEntry(add_item_frame, textvariable=self.app.invoice_rate_var, width=70)
        rate_entry.grid(row=1, column=7, padx=2, pady=6, sticky='ew')

        ctk.CTkLabel(add_item_frame, text="Total:", width=35, anchor='w').grid(row=1, column=8, padx=4, pady=6, sticky='w')
        self.app.invoice_item_total_var = tk.StringVar(value='0.00')
        total_label = ctk.CTkLabel(add_item_frame, textvariable=self.app.invoice_item_total_var,
                                   font=('Arial', 11, 'bold'), width=70)
        total_label.grid(row=1, column=9, padx=2, pady=6, sticky='ew')

        ctk.CTkButton(add_item_frame, text="Add", width=50, height=28, command=self.add_invoice_item).grid(row=1, column=10, padx=3, pady=6, sticky='ew')

        for col in [1, 3, 5, 7, 9]:
            add_item_frame.grid_columnconfigure(col, weight=1)

        tree_frame = ctk.CTkFrame(items_section, fg_color="transparent")
        tree_frame.pack(fill='both', expand=True, pady=(0, 12))

        from utils import make_treeview
        self.app.invoice_items_tree = make_treeview(
            tree_frame,
            columns=('Item', 'Quantity', 'Rate', 'Total'),
            headings=('Item Name', 'Quantity', 'Rate (PKR)', 'Total (PKR)'),
            widths=(280, 100, 100, 120),
            height=10
        )
        self.app.invoice_items_tree.bind('<Delete>', lambda e: self.delete_invoice_item())
        self.app.invoice_items_tree.bind('<BackSpace>', lambda e: self.delete_invoice_item())
        self.app.invoice_items_tree.bind('<Button-3>', lambda e: self.show_invoice_item_context_menu(e))
        self.app.invoice_items_tree.bind('<Button-2>', lambda e: self.show_invoice_item_context_menu(e))
        self.app.invoice_items_tree.bind('<Control-Button-1>', lambda e: self.show_invoice_item_context_menu(e))

        bottom_frame = ctk.CTkFrame(invoice_frame, fg_color="transparent")
        bottom_frame.pack(fill='x', padx=15, pady=(0, 20))

        left_btns = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        left_btns.pack(side='left')
        ctk.CTkButton(left_btns, text="🔄 Update to Today's Rates", width=160,
                      command=self.update_invoice_rates_from_current_list).pack(side='left', padx=5)

        right_section = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        right_section.pack(side='right')
        ctk.CTkLabel(right_section, text="Total:", font=('Arial', 12, 'bold')).pack(side='left', padx=10)
        self.app.invoice_total_var = tk.StringVar(value='PKR 0.00')
        ctk.CTkLabel(right_section, textvariable=self.app.invoice_total_var,
                     font=('Arial', 16, 'bold'), text_color=("#0066cc", "#3399ff")).pack(side='left', padx=5)
        ctk.CTkButton(right_section, text="📄 Generate Invoice", height=32,
                      command=self.generate_invoice).pack(side='left', padx=10)

        # RIGHT COLUMN: Recent Invoices
        right_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        right_frame.grid(row=1, column=1, sticky='nsew', padx=(5, 0))

        recent_frame = ctk.CTkFrame(right_frame, fg_color=("#ffffff", "#1a1a1a"))
        recent_frame.pack(fill='both', expand=True, padx=0, pady=0)

        recent_header_frame = ctk.CTkFrame(recent_frame, fg_color="transparent")
        recent_header_frame.pack(fill='x', padx=15, pady=(15, 12))
        ctk.CTkLabel(recent_header_frame, text="📋 Recent Invoices", font=('Arial', 14, 'bold')).pack(anchor='w', padx=0, pady=0)

        tree_frame2 = ctk.CTkFrame(recent_frame, fg_color="transparent")
        tree_frame2.pack(fill='both', expand=True, padx=15, pady=(0, 20))

        self.app.invoices_tree = make_treeview(
            tree_frame2,
            columns=('Invoice#', 'Customer'),
            headings=('Invoice #', 'Customer Name'),
            widths=(80, 120),
            height=20
        )

        # 🔥 macOS-Compatible Bindings
        self.app.invoices_tree.bind('<Double-Button-1>', self._on_invoice_double_click)
        self.app.invoices_tree.bind('<Button-3>', self.show_invoice_context_menu)
        self.app.invoices_tree.bind('<Button-2>', self.show_invoice_context_menu)
        self.app.invoices_tree.bind('<Control-Button-1>', self.show_invoice_context_menu)

        self.populate_invoices_tree()

    def populate_invoices_tree(self):
        """Populate the recent invoices tree view with invoices from the selected date"""
        if not hasattr(self.app, 'invoices_tree'):
            return
        for item in self.app.invoices_tree.get_children():
            self.app.invoices_tree.delete(item)
        selected_date = getattr(self.app, 'selected_date', '')
        if hasattr(self.app, 'invoices') and self.app.invoices:
            for invoice in reversed(self.app.invoices):
                if invoice.get('status') == 'deleted':
                    continue
                invoice_date = invoice.get('date', '')
                if not invoice_date and invoice.get('time'):
                    try:
                        from datetime import datetime
                        time_str = invoice.get('time', '')
                        date_part = time_str.split()[0] if time_str else ''
                        invoice_date_obj = datetime.strptime(date_part, "%d-%b-%Y")
                        invoice_date = invoice_date_obj.strftime("%Y-%m-%d")
                    except (ValueError, IndexError):
                        continue
                if invoice_date == selected_date:
                    self.app.invoices_tree.insert('', 'end', values=(
                        f"#{invoice.get('invoice_number', 'N/A')}",
                        invoice.get('customer_name', 'Unknown')
                    ))
