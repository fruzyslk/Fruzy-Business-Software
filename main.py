# main.py - Fruzy Business Manager with Calendar and Date-Based Data Persistence
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font as tkfont
from datetime import datetime, timedelta
import calendar as cal_module
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import json
import sys
import subprocess
import shutil
from pathlib import Path
from PIL import Image, ImageTk
# Import modular tabs
from purchase_entry import PurchaseEntryTab
from sales_entry import SalesEntryTab
from customer_invoice import CustomerInvoiceTab
from daily_summary import DailySummaryTab

# Set CustomTkinter appearance
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

class FruzyBusinessManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Fruzy - Vegetable & Fruit Business Manager")
        self.root.geometry("1400x900")
        # Fruzy color scheme
        self.colors = {
            'primary': '#2ecc71',
            'secondary': '#f39c12',
            'dark': '#27ae60',
            'light': '#ecf9f2',
            'accent': '#e67e22',
            'text_dark': '#2c3e50',
            'red': '#e74c3c',
            'white': '#ffffff'
        }
        # ✅ PORTABLE: All data stored in ./data/ under the script directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.data_dir = os.path.join(script_dir, 'data')
        self.invoices_dir = os.path.join(self.data_dir, 'Customer_Invoices')
        os.makedirs(self.invoices_dir, exist_ok=True)

        # Load data structures
        self.vegetables = self.load_vegetables()
        self.all_purchases = self.load_all_purchases()
        self.all_sales = self.load_all_sales()

        # Normalize data on load
        for date in self.all_purchases:
            self.all_purchases[date] = self.normalize_transaction_data(self.all_purchases[date], 'kg')
        for date in self.all_sales:
            self.all_sales[date] = self.normalize_transaction_data(self.all_sales[date], 'kg')

        # Current date tracking
        self.current_date = datetime.now()
        self.selected_date = self.current_date.strftime("%Y-%m-%d")

        # Load data for today
        self.purchases = self.get_purchases_for_date(self.selected_date)
        self.sales = self.get_sales_for_date(self.selected_date)
        self.purchases = self.normalize_transaction_data(self.purchases, 'kg')
        self.sales = self.normalize_transaction_data(self.sales, 'kg')

        # Invoice data
        self.invoices = self.load_invoices()
        self.rate_list = {}
        self.invoice_counter = self.load_invoice_counter()
        self.editing_invoice_number = None

        # UI variables
        self.purchase_veg_var = ctk.StringVar()
        self.purchase_qty_var = ctk.StringVar()
        self.purchase_rate_var = ctk.StringVar()
        self.purchase_unit_var = ctk.StringVar(value='kg')
        self.purchase_vendor_var = ctk.StringVar(value='Main Vendor')
        self.purchase_payment_var = ctk.StringVar(value='cash')
        self.purchase_total_var = ctk.StringVar(value='0.00')
        self.purchase_tree = None

        self.sales_veg_var = ctk.StringVar()
        self.sales_qty_var = ctk.StringVar()
        self.sales_rate_var = ctk.StringVar()
        self.sales_unit_var = ctk.StringVar(value='kg')
        self.sales_total_var = ctk.StringVar(value='0.00')
        self.sales_tree = None

        self.invoice_item_var = ctk.StringVar()
        self.invoice_qty_var = ctk.StringVar()
        self.invoice_rate_var = ctk.StringVar()
        self.invoice_unit_var = ctk.StringVar(value='kg')
        self.invoice_size_var = ctk.StringVar(value='Normal')
        self.invoice_item_total_var = ctk.StringVar(value='0.00')

        # Summary widgets
        self.total_purchase_label = None
        self.total_sales_label = None
        self.profit_label = None
        self.profit_percent_label = None
        self.purchase_items_label = None
        self.cash_purchase_label = None
        self.credit_purchase_label = None
        self.sales_items_label = None
        self.invoice_sales_label = None
        self.manual_sales_label = None
        self.avg_sale_label = None
        self.profit_tree = None
        self.qty_movement_tree = None
        self.current_date_label = None

        self.create_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    # ============ DATA NORMALIZATION ============
    def normalize_transaction_data(self, transactions, default_unit='kg'):
        """Ensure all transactions have 'quantity' as 'value unit' string and proper name fields."""
        normalized = []
        for t in transactions:
            t = t.copy()
            # Ensure quantity is formatted
            qty = t.get('quantity', '0')
            if isinstance(qty, (int, float)):
                t['quantity'] = f"{float(qty):.2f} {default_unit}"
            elif isinstance(qty, str) and qty.strip() and ' ' not in qty.strip():
                try:
                    float_val = float(qty.strip())
                    t['quantity'] = f"{float_val:.2f} {default_unit}"
                except ValueError:
                    t['quantity'] = f"0.00 {default_unit}"
            elif not qty.strip():
                t['quantity'] = f"0.00 {default_unit}"

            # Ensure name fields exist
            if 'vegetable_urdu' not in t or 'vegetable_english' not in t:
                veg_data = self.get_vegetable_data(t.get('vegetable', ''))
                if veg_data:
                    t['vegetable_urdu'] = veg_data['urdu']
                    t['vegetable_english'] = veg_data['english']
                else:
                    t['vegetable_urdu'] = t.get('vegetable', '')
                    t['vegetable_english'] = t.get('vegetable', '')

            # Ensure display name
            t['vegetable_display'] = f"{t['vegetable_urdu']} ({t['vegetable_english']})"
            normalized.append(t)
        return normalized

    # ============ DATA PERSISTENCE FUNCTIONS ============
    def load_vegetables(self):
        default_vegetables = [
            {"id": 1, "urdu": "ٹماٹر", "english": "Tomato"},
            {"id": 2, "urdu": "سبز مرچ", "english": "Green Chili"},
            {"id": 3, "urdu": "لہسن", "english": "Garlic"},
            {"id": 4, "urdu": "پیاز", "english": "Onion"},
            {"id": 5, "urdu": "آلو پرانا", "english": "Old Potato"}
        ]
        veg_path = os.path.join(self.data_dir, 'vegetables.json')
        if os.path.exists(veg_path):
            try:
                with open(veg_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading vegetables: {e}")
        self.save_vegetables(default_vegetables)
        return default_vegetables

    def save_vegetables(self, vegetables=None):
        if vegetables is None:
            vegetables = self.vegetables
        veg_path = os.path.join(self.data_dir, 'vegetables.json')
        try:
            with open(veg_path, 'w', encoding='utf-8') as f:
                json.dump(vegetables, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving vegetables: {e}")

    def load_all_purchases(self):
        purchase_path = os.path.join(self.data_dir, 'purchases_by_date.json')
        if os.path.exists(purchase_path):
            try:
                with open(purchase_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading purchases: {e}")
        return {}

    def save_all_purchases(self):
        purchase_path = os.path.join(self.data_dir, 'purchases_by_date.json')
        try:
            backup_path = purchase_path + '.backup'
            if os.path.exists(purchase_path):
                shutil.copy2(purchase_path, backup_path)
            temp_path = purchase_path + '.tmp'
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(self.all_purchases, f, ensure_ascii=False, indent=2)
            with open(temp_path, 'r', encoding='utf-8') as f:
                test_load = json.load(f)
            if os.path.exists(purchase_path):
                os.remove(purchase_path)
            os.rename(temp_path, purchase_path)
            print(f"✓ Purchases saved successfully ({len(self.all_purchases)} dates)")
        except Exception as e:
            print(f"✗ Error saving purchases: {e}")
            try:
                if os.path.exists(backup_path) and not os.path.exists(purchase_path):
                    shutil.copy2(backup_path, purchase_path)
                    print(f"Restored from backup")
            except:
                pass

    def load_all_sales(self):
        sales_path = os.path.join(self.data_dir, 'sales_by_date.json')
        if os.path.exists(sales_path):
            try:
                with open(sales_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading sales: {e}")
        return {}

    def save_all_sales(self):
        sales_path = os.path.join(self.data_dir, 'sales_by_date.json')
        try:
            backup_path = sales_path + '.backup'
            if os.path.exists(sales_path):
                shutil.copy2(sales_path, backup_path)
            temp_path = sales_path + '.tmp'
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(self.all_sales, f, ensure_ascii=False, indent=2)
            with open(temp_path, 'r', encoding='utf-8') as f:
                test_load = json.load(f)
            if os.path.exists(sales_path):
                os.remove(sales_path)
            os.rename(temp_path, sales_path)
            print(f"✓ Sales saved successfully ({len(self.all_sales)} dates)")
        except Exception as e:
            print(f"✗ Error saving sales: {e}")
            try:
                if os.path.exists(backup_path) and not os.path.exists(sales_path):
                    shutil.copy2(backup_path, sales_path)
                    print(f"Restored from backup")
            except:
                pass

    def get_purchases_for_date(self, date_str):
        return self.all_purchases.get(date_str, [])

    def get_sales_for_date(self, date_str):
        return self.all_sales.get(date_str, [])

    def save_purchases(self):
        try:
            self.all_purchases[self.selected_date] = self.purchases
            self.save_all_purchases()
        except Exception as e:
            print(f"Error in save_purchases: {e}")
            messagebox.showerror("Save Error", f"Failed to save purchases: {str(e)}")

    def save_sales(self):
        try:
            self.all_sales[self.selected_date] = self.sales
            self.save_all_sales()
        except Exception as e:
            print(f"Error in save_sales: {e}")
            messagebox.showerror("Save Error", f"Failed to save sales: {str(e)}")

    def set_date(self, date_str):
        self.all_purchases[self.selected_date] = self.purchases
        self.all_sales[self.selected_date] = self.sales
        self.save_all_purchases()
        self.save_all_sales()
        self.selected_date = date_str
        self.purchases = self.get_purchases_for_date(date_str)
        self.sales = self.get_sales_for_date(date_str)
        self.purchases = self.normalize_transaction_data(self.purchases, 'kg')
        self.sales = self.normalize_transaction_data(self.sales, 'kg')
        self.refresh_all_trees()
        self.update_summary()
        self.update_date_label()

    # ============ PUBLIC METHOD FOR TABS ============
    def refresh_daily_summary(self):
        if hasattr(self, 'summary_tab_instance'):
            self.summary_tab_instance.refresh_all_data()

    # ============ VEGETABLE NAME HELPER FUNCTIONS ============
    def get_vegetable_data(self, veg_display_name):
        if not veg_display_name:
            return None
        for veg in self.vegetables:
            display_text = f"{veg['urdu']} ({veg['english']})"
            if veg_display_name == display_text:
                return {'urdu': veg['urdu'], 'english': veg['english']}
        for veg in self.vegetables:
            if veg['english'].lower() == veg_display_name.lower():
                return {'urdu': veg['urdu'], 'english': veg['english']}
        for veg in self.vegetables:
            if veg['urdu'] == veg_display_name:
                return {'urdu': veg['urdu'], 'english': veg['english']}
        if '(' in veg_display_name and ')' in veg_display_name:
            try:
                urdu_part = veg_display_name.split('(')[0].strip()
                english_part = veg_display_name.split('(')[1].split(')')[0].strip()
                return {'urdu': urdu_part, 'english': english_part}
            except:
                pass
        return {'urdu': veg_display_name, 'english': veg_display_name}

    def get_vegetable_display_name(self, veg_data):
        if isinstance(veg_data, dict):
            urdu = veg_data.get('urdu', '')
            english = veg_data.get('english', '')
            if urdu and english:
                return f"{urdu} ({english})"
            elif urdu:
                return urdu
            elif english:
                return english
        if isinstance(veg_data, str):
            if '(' in veg_data and ')' in veg_data:
                return veg_data
            veg_info = self.get_vegetable_data(veg_data)
            if veg_info:
                return f"{veg_info['urdu']} ({veg_info['english']})"
            return veg_data
        return str(veg_data)

    # ============ INVOICE FUNCTIONS ============
    def load_invoice_counter(self):
        path = os.path.join(self.data_dir, 'invoice_counter.json')
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                    return data.get('last_invoice', 0)
            except:
                pass
        return 0

    def save_invoice_counter(self):
        path = os.path.join(self.data_dir, 'invoice_counter.json')
        try:
            with open(path, 'w') as f:
                json.dump({'last_invoice': self.invoice_counter}, f)
        except Exception as e:
            print(f"Error saving invoice counter: {e}")

    def load_invoices(self):
        path = os.path.join(self.data_dir, 'invoices.json')
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return []

    def save_invoices(self):
        path = os.path.join(self.data_dir, 'invoices.json')
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(self.invoices, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving invoices: {e}")

    # ============ UI CREATION ============
    def create_ui(self):
        header_frame = ctk.CTkFrame(self.root, height=80, corner_radius=0)
        header_frame.pack(fill='x', padx=0, pady=0)
        header_frame.pack_propagate(False)

        logo_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        logo_frame.pack(side='left', padx=20, pady=10)
        self.logo_image = None
        try:
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fruzy_logo.png')
            if os.path.exists(logo_path):
                img = Image.open(logo_path)
                img = img.resize((60, 60), Image.Resampling.LANCZOS)
                self.logo_image = ImageTk.PhotoImage(img)
                logo_label = ctk.CTkLabel(logo_frame, image=self.logo_image, text="")
                logo_label.pack(side='left', padx=(0, 15))
        except Exception as e:
            print(f"Warning: Could not load logo: {e}")

        ctk.CTkLabel(logo_frame, text=" FRUZY", font=('Arial', 28, 'bold')).pack(side='left')
        ctk.CTkLabel(logo_frame, text="Business Manager", font=('Arial', 10)).pack()

        self.current_date_label = ctk.CTkLabel(
            header_frame,
            text=self.current_date.strftime("%A, %B %d, %Y"),
            font=('Arial', 14, 'bold')
        )
        self.current_date_label.pack(side='left', padx=20, pady=15)

        btn_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        btn_frame.pack(side='right', padx=20, pady=15)
        ctk.CTkButton(btn_frame, text="📅 Calendar", font=('Arial', 11, 'bold'),
                      command=self.open_calendar_dialog).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="📥 Import Data", font=('Arial', 11, 'bold'),
                      command=self.import_excel_data).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="📂 Open Invoices", font=('Arial', 11, 'bold'),
                      command=self.open_invoice_folder).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="📊 Export", font=('Arial', 11, 'bold'),
                      command=self.export_to_excel).pack(side='left', padx=5)

        main_container = ctk.CTkFrame(self.root)
        main_container.pack(fill='both', expand=True, padx=10, pady=10)
        self.create_sidebar(main_container)
        self.create_notebook(main_container)

    def create_sidebar(self, parent):
        sidebar = ctk.CTkFrame(parent, width=320)
        sidebar.pack(side='left', fill='y', padx=(0, 10))
        sidebar.pack_propagate(False)

        sidebar_header = ctk.CTkFrame(sidebar, height=60)
        sidebar_header.pack(fill='x')
        sidebar_header.pack_propagate(False)
        ctk.CTkLabel(sidebar_header, text="🥬 Vegetables & Fruits", font=('Arial', 16, 'bold')).pack(pady=15)

        search_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        search_frame.pack(fill='x', padx=10, pady=10)
        ctk.CTkLabel(search_frame, text="🔍", font=('Arial', 13)).pack(side='left', padx=5)
        self.search_var = ctk.StringVar()
        self.search_var.trace('w', self.filter_vegetables)
        ctk.CTkEntry(search_frame, textvariable=self.search_var, font=('Arial', 10), width=180).pack(side='left', fill='x', expand=True)

        ctk.CTkButton(sidebar, text="➕ Add New Item", font=('Arial', 10, 'bold'),
                      command=self.add_vegetable_dialog).pack(fill='x', padx=10, pady=5)

        list_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')

        urdu_font = ('Jameel Noori Nastaleeq', 16)
        try:
            tkfont.Font(font=urdu_font)
            actual_font = urdu_font
        except:
            actual_font = ('Arial', 14)

        self.veg_listbox = tk.Listbox(list_frame, font=actual_font,
                                      yscrollcommand=scrollbar.set,
                                      selectbackground=self.colors['primary'],
                                      selectforeground='white',
                                      relief='solid', bd=1, height=20)
        self.veg_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.veg_listbox.yview)
        self.veg_listbox.bind('<Double-Button-1>', self.select_vegetable)
        self.veg_listbox.bind('<Button-3>', self.show_veg_context_menu)
        self.populate_vegetable_list()

        btn_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        btn_frame.pack(fill='x', padx=10, pady=5)
        ctk.CTkButton(btn_frame, text="✏️ Edit", font=('Arial', 9, 'bold'),
                      command=self.edit_vegetable_dialog).pack(side='left', fill='x', expand=True, padx=2)
        ctk.CTkButton(btn_frame, text="🗑️ Delete", font=('Arial', 9, 'bold'),
                      command=self.delete_vegetable).pack(side='left', fill='x', expand=True, padx=2)

    def populate_vegetable_list(self, filter_text=''):
        self.veg_listbox.delete(0, tk.END)
        for veg in self.vegetables:
            if (filter_text.lower() in veg['english'].lower() or filter_text in veg['urdu']):
                display_text = f"{veg['urdu']} ({veg['english']})"
                self.veg_listbox.insert(tk.END, display_text)

    def filter_vegetables(self, *args):
        self.populate_vegetable_list(self.search_var.get())

    def select_vegetable(self, event):
        """Select vegetable and populate current tab (Purchase, Sales, or Invoice)"""
        selection = self.veg_listbox.curselection()
        if selection:
            veg_name = self.veg_listbox.get(selection[0])
            current_tab = self.notebook.index(self.notebook.select())
            if current_tab == 0:
                self.purchase_veg_var.set(veg_name)
            elif current_tab == 1:
                self.sales_veg_var.set(veg_name)
            elif current_tab == 2:
                self.invoice_item_var.set(veg_name)
            if hasattr(self, 'invoice_tab_instance'):
                self.invoice_tab_instance.load_item_rate()

    def add_vegetable_dialog(self):
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Add New Item - Fruzy")
        dialog.geometry("500x280")
        dialog.transient(self.root)
        dialog.grab_set()
        ctk.CTkLabel(dialog, text="Add New Item", font=('Arial', 14, 'bold')).pack(pady=15)
        frame = ctk.CTkFrame(dialog)
        frame.pack(padx=30, pady=15, fill='both', expand=True)

        ctk.CTkLabel(frame, text="Urdu Name:", font=('Arial', 11, 'bold')).grid(row=0, column=0, sticky='w', pady=10)
        urdu_entry = ctk.CTkEntry(frame, font=('Arial', 12), width=350)
        urdu_entry.grid(row=0, column=1, padx=15, pady=10, sticky='ew')

        ctk.CTkLabel(frame, text="English Name:", font=('Arial', 11, 'bold')).grid(row=1, column=0, sticky='w', pady=10)
        english_entry = ctk.CTkEntry(frame, font=('Arial', 12), width=350)
        english_entry.grid(row=1, column=1, padx=15, pady=10, sticky='ew')
        frame.grid_columnconfigure(1, weight=1)

        def save():
            urdu = urdu_entry.get().strip()
            english = english_entry.get().strip()
            if not urdu or not english:
                messagebox.showwarning("Missing Data", "Please fill both fields")
                return
            new_id = max([v['id'] for v in self.vegetables], default=0) + 1
            self.vegetables.append({'id': new_id, 'urdu': urdu, 'english': english})
            self.save_vegetables()
            self.populate_vegetable_list()
            messagebox.showinfo("Success", "Item added!")
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog)
        btn_frame.pack(pady=15)
        ctk.CTkButton(btn_frame, text="Save", command=save, width=100).pack(side='left', padx=10)
        ctk.CTkButton(btn_frame, text="Cancel", command=dialog.destroy, width=100).pack(side='left', padx=10)

    def edit_vegetable_dialog(self):
        selection = self.veg_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an item")
            return
        selected_text = self.veg_listbox.get(selection[0])
        index_in_full_list = None
        for i, veg in enumerate(self.vegetables):
            if f"{veg['urdu']} ({veg['english']})" == selected_text:
                index_in_full_list = i
                break
        if index_in_full_list is None:
            return
        veg = self.vegetables[index_in_full_list]

        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Edit Item - Fruzy")
        dialog.geometry("500x280")
        dialog.transient(self.root)
        dialog.grab_set()
        ctk.CTkLabel(dialog, text="Edit Item", font=('Arial', 14, 'bold')).pack(pady=15)
        frame = ctk.CTkFrame(dialog)
        frame.pack(padx=30, pady=15, fill='both', expand=True)

        ctk.CTkLabel(frame, text="Urdu Name:", font=('Arial', 11, 'bold')).grid(row=0, column=0, sticky='w', pady=10)
        urdu_entry = ctk.CTkEntry(frame, font=('Arial', 12), width=350)
        urdu_entry.insert(0, veg['urdu'])
        urdu_entry.grid(row=0, column=1, padx=15, pady=10, sticky='ew')

        ctk.CTkLabel(frame, text="English Name:", font=('Arial', 11, 'bold')).grid(row=1, column=0, sticky='w', pady=10)
        english_entry = ctk.CTkEntry(frame, font=('Arial', 12), width=350)
        english_entry.insert(0, veg['english'])
        english_entry.grid(row=1, column=1, padx=15, pady=10, sticky='ew')
        frame.grid_columnconfigure(1, weight=1)

        def update():
            urdu = urdu_entry.get().strip()
            english = english_entry.get().strip()
            if not urdu or not english:
                messagebox.showwarning("Missing Data", "Please fill both fields")
                return
            if index_in_full_list < 0 or index_in_full_list >= len(self.vegetables):
                messagebox.showerror("Error", "Item no longer exists")
                dialog.destroy()
                return
            self.vegetables[index_in_full_list]['urdu'] = urdu
            self.vegetables[index_in_full_list]['english'] = english
            self.save_vegetables()
            self.populate_vegetable_list()
            messagebox.showinfo("Success", "Item updated!")
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog)
        btn_frame.pack(pady=15)
        ctk.CTkButton(btn_frame, text="Update", command=update, width=100).pack(side='left', padx=10)
        ctk.CTkButton(btn_frame, text="Cancel", command=dialog.destroy, width=100).pack(side='left', padx=10)

    def delete_vegetable(self):
        selection = self.veg_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an item")
            return
        if not messagebox.askyesno("Confirm", "Delete this item?"):
            return
        selected_text = self.veg_listbox.get(selection[0])
        veg_to_delete = None
        for veg in self.vegetables:
            if f"{veg['urdu']} ({veg['english']})" == selected_text:
                veg_to_delete = veg
                break
        if veg_to_delete:
            self.vegetables = [v for v in self.vegetables if v['id'] != veg_to_delete['id']]
            self.save_vegetables()
            self.populate_vegetable_list()
            messagebox.showinfo("Success", "Item deleted!")

    def show_veg_context_menu(self, event):
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Edit", command=self.edit_vegetable_dialog)
        menu.add_command(label="Delete", command=self.delete_vegetable)
        menu.post(event.x_root, event.y_root)

    # ============ NOTEBOOK (TABS) ============
    def create_notebook(self, parent):
        notebook_frame = ctk.CTkFrame(parent)
        notebook_frame.pack(side='left', fill='both', expand=True)
        self.notebook = ttk.Notebook(notebook_frame)
        self.notebook.pack(fill='both', expand=True)

        self.purchase_tab = ctk.CTkFrame(self.notebook, fg_color=self.colors['light'])
        self.notebook.add(self.purchase_tab, text='📦 Purchase Entry')
        self.purchase_tab_instance = PurchaseEntryTab(self.purchase_tab, self)

        self.sales_tab = ctk.CTkFrame(self.notebook, fg_color=self.colors['light'])
        self.notebook.add(self.sales_tab, text='💰 Sales Entry')
        self.sales_tab_instance = SalesEntryTab(self.sales_tab, self)

        self.invoice_tab = ctk.CTkFrame(self.notebook, fg_color=self.colors['light'])
        self.notebook.add(self.invoice_tab, text='🧾 Customer Invoice')
        self.invoice_tab_instance = CustomerInvoiceTab(self.invoice_tab, self)

        self.summary_tab = ctk.CTkFrame(self.notebook, fg_color=self.colors['light'])
        self.notebook.add(self.summary_tab, text='📊 Daily Summary')
        self.summary_tab_instance = DailySummaryTab(self.summary_tab, self)

    # ============ CALENDAR FEATURE ============
    def open_calendar_dialog(self):
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Select Date - Fruzy")
        dialog.geometry("380x500")
        dialog.transient(self.root)
        dialog.grab_set()
        ctk.CTkLabel(dialog, text="📅 Select Date", font=('Arial', 16, 'bold')).pack(pady=15)
        today = self.current_date
        year = today.year
        month = today.month

        nav_top_frame = ctk.CTkFrame(dialog)
        nav_top_frame.pack(fill='x', padx=20, pady=10)
        month_year_label = ctk.CTkLabel(nav_top_frame, text=f"{cal_module.month_name[month]} {year}", font=('Arial', 14, 'bold'))
        month_year_label.pack(pady=5)

        button_nav_frame = ctk.CTkFrame(nav_top_frame)
        button_nav_frame.pack(fill='x', pady=10)

        def show_calendar(new_year, new_month):
            for widget in cal_frame.winfo_children():
                widget.destroy()
            days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            for col, day in enumerate(days):
                ctk.CTkLabel(cal_frame, text=day, font=('Arial', 10, 'bold')).grid(row=0, column=col, padx=5, pady=5)
            cal_obj = cal_module.monthcalendar(new_year, new_month)
            for row, week in enumerate(cal_obj, 1):
                for col, day in enumerate(week):
                    if day == 0:
                        continue
                    date_str = f"{new_year:04d}-{new_month:02d}-{day:02d}"
                    has_data = date_str in self.all_purchases or date_str in self.all_sales
                    bg_color = self.colors['primary'] if has_data else self.colors['light']
                    btn = ctk.CTkButton(
                        cal_frame,
                        text=str(day),
                        width=40,
                        height=40,
                        fg_color=bg_color,
                        command=lambda d=day, m=new_month, y=new_year: self.select_calendar_date(d, m, y, dialog)
                    )
                    btn.grid(row=row, column=col, padx=2, pady=2)
            month_year_label.configure(text=f"{cal_module.month_name[new_month]} {new_year}")

        def prev_month():
            nonlocal year, month
            month -= 1
            if month < 1:
                month = 12
                year -= 1
            show_calendar(year, month)

        def next_month():
            nonlocal year, month
            month += 1
            if month > 12:
                month = 1
                year += 1
            show_calendar(year, month)

        ctk.CTkButton(button_nav_frame, text="◀ Prev", width=60, command=prev_month).pack(side='left', padx=5)
        ctk.CTkButton(button_nav_frame, text="Today", width=60,
                      command=lambda: self.set_date(datetime.now().strftime("%Y-%m-%d"))).pack(side='left', padx=5)
        ctk.CTkButton(button_nav_frame, text="Next ▶", width=60, command=next_month).pack(side='left', padx=5)

        cal_frame = ctk.CTkFrame(dialog)
        cal_frame.pack(padx=20, pady=10, fill='both', expand=True)
        show_calendar(year, month)

        nav_btn_frame = ctk.CTkFrame(dialog)
        nav_btn_frame.pack(pady=15)
        ctk.CTkButton(nav_btn_frame, text="Close", command=dialog.destroy).pack(side='left', padx=5)

    def select_calendar_date(self, day, month, year, dialog):
        date_str = f"{year:04d}-{month:02d}-{day:02d}"
        self.set_date(date_str)
        dialog.destroy()

    def update_date_label(self):
        try:
            date_obj = datetime.strptime(self.selected_date, "%Y-%m-%d")
            self.current_date_label.configure(text=date_obj.strftime("%A, %B %d, %Y"))
        except:
            pass

    # ============ TRANSACTION FUNCTIONS ============
    def calculate_purchase_total(self, *args):
        try:
            qty = float(self.purchase_qty_var.get()) if self.purchase_qty_var.get() else 0
            rate = float(self.purchase_rate_var.get()) if self.purchase_rate_var.get() else 0
            self.purchase_total_var.set(f"{qty * rate:.2f}")
        except:
            self.purchase_total_var.set("0.00")

    def calculate_sales_total(self, *args):
        try:
            qty = float(self.sales_qty_var.get()) if self.sales_qty_var.get() else 0
            rate = float(self.sales_rate_var.get()) if self.sales_rate_var.get() else 0
            self.sales_total_var.set(f"{qty * rate:.2f}")
        except:
            self.sales_total_var.set("0.00")

    def add_purchase(self):
        if not all([self.purchase_veg_var.get(), self.purchase_qty_var.get(), self.purchase_rate_var.get()]):
            messagebox.showwarning("Missing Data", "Please fill all fields")
            return
        try:
            qty = float(self.purchase_qty_var.get())
            if qty <= 0:
                messagebox.showerror("Invalid Data", "Quantity must be greater than 0")
                return
            rate = float(self.purchase_rate_var.get())
            if rate < 0:
                messagebox.showerror("Invalid Data", "Rate cannot be negative")
                return
            total = qty * rate
            veg_data = self.get_vegetable_data(self.purchase_veg_var.get())
            if veg_data is None:
                veg_data = {'urdu': self.purchase_veg_var.get(), 'english': self.purchase_veg_var.get()}
            purchase = {
                'vegetable_urdu': veg_data['urdu'],
                'vegetable_english': veg_data['english'],
                'vegetable_display': f"{veg_data['urdu']} ({veg_data['english']})",
                'quantity': f"{qty:.2f} {self.purchase_unit_var.get()}",
                'rate': f"{rate:.2f}",
                'total': f"{total:.2f}",
                'vendor': self.purchase_vendor_var.get(),
                'payment': self.purchase_payment_var.get()
            }
            self.purchases.append(purchase)
            self.purchase_veg_var.set('')
            self.purchase_qty_var.set('')
            self.purchase_rate_var.set('')
            self.purchase_total_var.set('0.00')
            self.save_purchases()
            self.update_summary()
            messagebox.showinfo("Success", "Purchase added!")
        except ValueError:
            messagebox.showerror("Invalid Data", "Please enter valid numbers")

    def add_sale(self):
        if not all([self.sales_veg_var.get(), self.sales_qty_var.get(), self.sales_rate_var.get()]):
            messagebox.showwarning("Missing Data", "Please fill all fields")
            return
        try:
            qty = float(self.sales_qty_var.get())
            if qty <= 0:
                messagebox.showerror("Invalid Data", "Quantity must be greater than 0")
                return
            rate = float(self.sales_rate_var.get())
            if rate < 0:
                messagebox.showerror("Invalid Data", "Rate cannot be negative")
                return
            total = qty * rate
            veg_data = self.get_vegetable_data(self.sales_veg_var.get())
            if veg_data is None:
                veg_data = {'urdu': self.sales_veg_var.get(), 'english': self.sales_veg_var.get()}
            sale = {
                'source': 'Manual Entry',
                'vegetable_urdu': veg_data['urdu'],
                'vegetable_english': veg_data['english'],
                'vegetable_display': f"{veg_data['urdu']} ({veg_data['english']})",
                'quantity': f"{qty:.2f} {self.sales_unit_var.get()}",
                'rate': f"{rate:.2f}",
                'total': f"{total:.2f}"
            }
            self.sales.append(sale)
            self.sales_veg_var.set('')
            self.sales_qty_var.set('')
            self.sales_rate_var.set('')
            self.sales_total_var.set('0.00')
            self.save_sales()
            self.update_summary()
            self.sales_tab_instance.reload_sales_list()
            messagebox.showinfo("Success", "Sale added!")
        except ValueError:
            messagebox.showerror("Invalid Data", "Please enter valid numbers")

    def delete_purchase(self):
        if not self.purchase_tree:
            messagebox.showwarning("No Data", "No purchases to delete")
            return
        selection = self.purchase_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a purchase")
            return
        if messagebox.askyesno("Confirm", "Delete this purchase?"):
            index = self.purchase_tree.index(selection[0])
            self.purchases.pop(index)
            self.purchase_tree.delete(selection)
            self.save_purchases()
            self.update_summary()
    def delete_sale(self):
        """Delete selected sales entries (supports multi-select)."""
        selected = self.sales_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select one or more rows to delete.")
            return

        if not messagebox.askyesno("Confirm Delete", f"Delete {len(selected)} selected sale(s)?"):
            return

        # Try to interpret iids as indices
        try:
            indices = sorted([int(iid) for iid in selected], reverse=True)
        except ValueError:
            # Invalid iids found → likely from refresh_sales_tree
            messagebox.showerror(
                "Sync Error",
                "Sales list is out of sync. Reloading...\nPlease try again."
            )
            if hasattr(self, 'sales_tab_instance'):
                self.sales_tab_instance.reload_sales_list()
            return

        # Delete from data list
        for idx in indices:
            if 0 <= idx < len(self.sales):
                del self.sales[idx]

        # Save and refresh via the tab (which uses correct iids)
        self.save_sales()
        self.update_summary()
        self.sales_tab_instance.reload_sales_list()
    def update_summary(self):
        total_purchase = sum(float(p['total']) for p in self.purchases) if self.purchases else 0
        total_sales = sum(float(s['total']) for s in self.sales) if self.sales else 0
        profit = total_sales - total_purchase
        profit_percent = (profit / total_purchase * 100) if total_purchase > 0 else 0

        if self.total_purchase_label:
            self.total_purchase_label.configure(text=f"PKR {total_purchase:,.2f}")
        if self.total_sales_label:
            self.total_sales_label.configure(text=f"PKR {total_sales:,.2f}")
        if self.profit_label:
            self.profit_label.configure(text=f"PKR {profit:,.2f}",
                                        text_color=self.colors['primary'] if profit >= 0 else self.colors['red'])
        if self.profit_percent_label:
            self.profit_percent_label.configure(text=f"({profit_percent:.2f}%)")
        if self.purchase_items_label:
            self.purchase_items_label.configure(text=str(len(self.purchases)))
        if self.sales_items_label:
            self.sales_items_label.configure(text=str(len(self.sales)))

        if hasattr(self, 'summary_tab_instance') and self.summary_tab_instance:
            try:
                self.summary_tab_instance.refresh_all_data()
            except Exception as e:
                print(f"Error refreshing daily summary: {e}")

    def refresh_purchase_tree(self):
        if self.purchase_tree:
            try:
                for item in self.purchase_tree.get_children():
                    self.purchase_tree.delete(item)
                for purchase in self.purchases:
                    self.purchase_tree.insert('', 'end', values=(
                        purchase['vegetable_display'],
                        purchase['quantity'],
                        purchase['rate'],
                        purchase['total'],
                        purchase['vendor'],
                        purchase['payment']
                    ))
            except:
                pass
    def refresh_all_trees(self):
        self.refresh_purchase_tree()
        self.sales_tab_instance.reload_sales_list()  # ← GOOD
        if hasattr(self, 'invoice_tab_instance'):
            self.invoice_tab_instance.populate_invoices_tree()

    def open_invoice_folder(self):
        try:
            invoices_folder = self.invoices_dir
            if sys.platform == "darwin":
                subprocess.run(["open", invoices_folder], check=False)
            elif sys.platform == "win32":
                os.startfile(invoices_folder)
            else:
                subprocess.run(["xdg-open", invoices_folder], check=False)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open folder: {str(e)}")

    def import_excel_data(self):
        filename = filedialog.askopenfilename(
            title="Select Excel File with Monthly Data",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not filename:
            return
        try:
            wb = openpyxl.load_workbook(filename)
            sheet_names = [s for s in wb.sheetnames if s != "Monthly Summary"]
            if not sheet_names:
                messagebox.showwarning("No Data", "No daily sheets found in this file")
                return

            imported_count = 0
            import_errors = []
            for sheet_name in sheet_names:
                try:
                    ws = wb[sheet_name]
                    date_str = self._parse_date_from_sheet_name(sheet_name)
                    if not date_str:
                        import_errors.append(f"Sheet '{sheet_name}': Could not parse date")
                        continue

                    purchases = []
                    sales = []
                    in_purchases = False
                    in_sales = False

                    for row in ws.iter_rows(min_row=1):
                        if not row or not row[0].value:
                            continue
                        if row[0].value == "PURCHASES":
                            in_purchases = True
                            in_sales = False
                            continue
                        elif row[0].value == "SALES":
                            in_sales = True
                            in_purchases = False
                            continue
                        elif row[0].value == "DAILY SUMMARY":
                            break

                        if in_purchases and row[0].value and row[0].value not in ["Vegetable", "PURCHASES", "Vegetable (اردو/English)"]:
                            try:
                                veg = str(row[0].value) if row[0].value else ""
                                qty = str(row[1].value) if len(row) > 1 and row[1].value else "0"
                                rate = str(row[2].value) if len(row) > 2 and row[2].value else "0"
                                total = str(row[3].value) if len(row) > 3 and row[3].value else "0"
                                vendor = str(row[4].value) if len(row) > 4 and row[4].value else "Main Vendor"
                                payment = str(row[5].value).lower() if len(row) > 5 and row[5].value else "cash"
                                if veg and total and float(total) > 0:
                                    veg_data = self.get_vegetable_data(veg)
                                    if veg_data is None:
                                        veg_data = {'urdu': veg, 'english': veg}
                                    if ' ' not in qty:
                                        qty = f"{qty} kg"
                                    purchase = {
                                        'vegetable_urdu': veg_data['urdu'],
                                        'vegetable_english': veg_data['english'],
                                        'vegetable_display': f"{veg_data['urdu']} ({veg_data['english']})",
                                        'quantity': qty,
                                        'rate': rate,
                                        'total': total,
                                        'vendor': vendor,
                                        'payment': payment
                                    }
                                    purchases.append(purchase)
                            except (ValueError, IndexError):
                                continue

                        if in_sales and row[0].value and row[0].value not in ["Vegetable", "SALES", "Vegetable (اردو/English)"]:
                            try:
                                veg = str(row[0].value) if row[0].value else ""
                                qty = str(row[1].value) if len(row) > 1 and row[1].value else "0"
                                rate = str(row[2].value) if len(row) > 2 and row[2].value else "0"
                                total = str(row[3].value) if len(row) > 3 and row[3].value else "0"
                                source = str(row[6].value) if len(row) > 6 and row[6].value else "Manual Entry"
                                if veg and total and float(total) > 0:
                                    veg_data = self.get_vegetable_data(veg)
                                    if veg_data is None:
                                        veg_data = {'urdu': veg, 'english': veg}
                                    if ' ' not in qty:
                                        qty = f"{qty} kg"
                                    sale = {
                                        'source': source,
                                        'vegetable_urdu': veg_data['urdu'],
                                        'vegetable_english': veg_data['english'],
                                        'vegetable_display': f"{veg_data['urdu']} ({veg_data['english']})",
                                        'quantity': qty,
                                        'rate': rate,
                                        'total': total
                                    }
                                    sales.append(sale)
                            except (ValueError, IndexError):
                                continue

                    if purchases or sales:
                        self.all_purchases[date_str] = purchases
                        existing_sales = self.all_sales.get(date_str, [])
                        invoice_sales = [s for s in existing_sales if 'invoice' in s.get('source', '').lower()]
                        self.all_sales[date_str] = sales + invoice_sales
                        imported_count += 1
                except Exception as e:
                    import_errors.append(f"Sheet '{sheet_name}': {str(e)}")
                    continue

            if imported_count > 0:
                self.save_all_purchases()
                self.save_all_sales()
                first_date = self._parse_date_from_sheet_name(sheet_names[0])
                if first_date:
                    self.set_date(first_date)

                error_info = ""
                if import_errors:
                    error_info = f"\nWarnings:\n" + "\n".join(import_errors[:5])
                    if len(import_errors) > 5:
                        error_info += f"\n... and {len(import_errors) - 5} more warnings"
                messagebox.showinfo("Success", f"✅ Imported data for {imported_count} dates!\nNavigate using the calendar to view different dates.{error_info}")
            else:
                error_msg = "No valid data was imported.\nMake sure your Excel file has:\n"
                error_msg += "- Sheet names as dates (DD-MMM-YYYY or YYYY-MM-DD)\n"
                error_msg += "- 'PURCHASES' and 'SALES' sections\n"
                error_msg += "- Data rows with at least Vegetable, Quantity, Rate, Total"
                if import_errors:
                    error_msg += "\nErrors:\n" + "\n".join(import_errors[:3])
                messagebox.showwarning("Import Failed", error_msg)
        except Exception as e:
            messagebox.showerror("Error", f"Import error: {str(e)}")

    def _parse_date_from_sheet_name(self, sheet_name):
        formats = ["%Y-%m-%d", "%d-%b-%Y", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]
        for fmt in formats:
            try:
                date_obj = datetime.strptime(sheet_name, fmt)
                return date_obj.strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    def export_to_excel(self):
        if not self.purchases and not self.sales:
            messagebox.showwarning("No Data", "Please add data first")
            return
        try:
            selected_date_obj = datetime.strptime(self.selected_date, "%Y-%m-%d")
            month_year = selected_date_obj.strftime("%B_%Y")
            filename = f"{month_year}.xlsx"
            if os.path.exists(filename):
                wb = openpyxl.load_workbook(filename)
            else:
                wb = openpyxl.Workbook()
                if wb.active:
                    wb.remove(wb.active)
                summary_sheet = wb.create_sheet("Monthly Summary", 0)
                self.create_summary_sheet(summary_sheet)
            sheet_name = selected_date_obj.strftime("%d-%b-%Y")
            if sheet_name in wb.sheetnames:
                if not messagebox.askyesno("Exists", f"Overwrite '{sheet_name}'?"):
                    return
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            self.write_daily_sheet(ws)
            if "Monthly Summary" in wb.sheetnames:
                self.update_summary_sheet(wb["Monthly Summary"], wb)
            wb.save(filename)
            messagebox.showinfo("Success", f"Exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Export error: {str(e)}")

    def write_daily_sheet(self, ws):
        header_fill = PatternFill(start_color=self.colors['dark'].replace('#', ''),
                                 end_color=self.colors['dark'].replace('#', ''), fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        ws['A1'] = "FRUZY - Vegetable & Fruit Business"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws.merge_cells('A1:F1')
        selected_date_obj = datetime.strptime(self.selected_date, "%Y-%m-%d")
        ws['A2'] = f"Date: {selected_date_obj.strftime('%A, %B %d, %Y')}"
        ws['A2'].font = Font(name='Arial', size=12)
        ws.merge_cells('A2:F2')
        row = 4
        ws[f'A{row}'] = "PURCHASES"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        row += 1
        for col, header in enumerate(['Vegetable (اردو/English)', 'Quantity', 'Rate (PKR)', 'Total (PKR)', 'Vendor', 'Payment Type'], 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        for purchase in self.purchases:
            # Get the display name with both Urdu and English
            veg_display = purchase.get('vegetable', purchase.get('vegetable_english', ''))
            if purchase.get('vegetable_urdu') and purchase.get('vegetable_english'):
                veg_display = f"{purchase.get('vegetable_urdu')} ({purchase.get('vegetable_english')})"
            
            ws.cell(row, 1, veg_display).border = border
            ws.cell(row, 2, purchase['quantity']).border = border
            try:
                ws.cell(row, 3, float(purchase['rate'])).border = border
            except Exception:
                ws.cell(row, 3, purchase['rate']).border = border
            ws.cell(row, 3).number_format = '#,##0.00'
            try:
                ws.cell(row, 4, float(purchase['total'])).border = border
            except Exception:
                ws.cell(row, 4, purchase['total']).border = border
            ws.cell(row, 4).number_format = '#,##0.00'
            ws.cell(row, 5, purchase['vendor']).border = border
            ws.cell(row, 6, purchase['payment'].upper()).border = border
            row += 1
        row += 2
        ws[f'A{row}'] = "SALES"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        row += 1
        for col, header in enumerate(['Vegetable (اردو/English)', 'Quantity', 'Rate (PKR)', 'Total (PKR)'], 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        for sale in self.sales:
            # Get the display name with both Urdu and English
            veg_display = sale.get('vegetable', sale.get('vegetable_english', ''))
            if sale.get('vegetable_urdu') and sale.get('vegetable_english'):
                veg_display = f"{sale.get('vegetable_urdu')} ({sale.get('vegetable_english')})"
            
            ws.cell(row, 1, veg_display).border = border
            ws.cell(row, 2, sale['quantity']).border = border
            try:
                ws.cell(row, 3, float(sale['rate'])).border = border
            except Exception:
                ws.cell(row, 3, sale['rate']).border = border
            ws.cell(row, 3).number_format = '#,##0.00'
            try:
                ws.cell(row, 4, float(sale['total'])).border = border
            except Exception:
                ws.cell(row, 4, sale['total']).border = border
            ws.cell(row, 4).number_format = '#,##0.00'
            row += 1
        total_purchase = sum(float(p['total']) for p in self.purchases) if self.purchases else 0
        total_sales = sum(float(s['total']) for s in self.sales) if self.sales else 0
        profit = total_sales - total_purchase
        profit_percent = (profit / total_purchase * 100) if total_purchase > 0 else 0
        row += 2
        ws[f'A{row}'] = "DAILY SUMMARY"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        row += 1
        ws.cell(row, 1, "Total Purchase:").font = Font(bold=True)
        ws.cell(row, 2, total_purchase).number_format = '"PKR "#,##0.00'
        ws.cell(row, 2).font = Font(bold=True)
        row += 1
        ws.cell(row, 1, "Total Sales:").font = Font(bold=True)
        ws.cell(row, 2, total_sales).number_format = '"PKR "#,##0.00'
        ws.cell(row, 2).font = Font(bold=True)
        row += 1
        ws.cell(row, 1, "Profit/Loss:").font = Font(bold=True)
        ws.cell(row, 2, profit).number_format = '"PKR "#,##0.00'
        ws.cell(row, 2).font = Font(bold=True)
        row += 1
        ws.cell(row, 1, "Profit Percentage:").font = Font(bold=True)
        ws.cell(row, 2, f"{profit_percent:.2f}%")
        ws.cell(row, 2).font = Font(bold=True)
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        try:
            totals = self.get_qty_totals()
            if totals:
                row += 2
                ws[f'A{row}'] = "QUANTITY MOVEMENT"
                ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
                row += 1
                headers = ['Vegetable', 'Purchased', 'Sold']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row, col_idx, header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                row += 1
                for veg_dict in self.vegetables:
                    veg_label = f"{veg_dict.get('urdu', '')} ({veg_dict.get('english', '')})"
                    entry = None
                    for key in totals.keys():
                        if veg_dict.get('english') and veg_dict.get('english') in key:
                            entry = totals[key]
                            break
                        if veg_dict.get('urdu') and veg_dict.get('urdu') in key:
                            entry = totals[key]
                            break
                        if key == veg_dict.get('english') or key == veg_dict.get('urdu'):
                            entry = totals[key]
                            break
                    if not entry:
                        continue
                    p_display = self._format_qty_buckets(entry.get('purchased', {}))
                    s_display = self._format_qty_buckets(entry.get('sold', {}))
                    if p_display == "0.00" and s_display == "0.00":
                        continue
                    ws.cell(row, 1, veg_label).border = border
                    ws.cell(row, 2, p_display).border = border
                    ws.cell(row, 3, s_display).border = border
                    row += 1
                for col in ['G', 'H', 'I']:
                    try:
                        ws.column_dimensions[col].width = 20
                    except Exception:
                        pass
        except Exception:
            pass

    def create_summary_sheet(self, ws):
        ws['A1'] = "FRUZY - Monthly Summary"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws.merge_cells('A1:E1')
        selected_date_obj = datetime.strptime(self.selected_date, "%Y-%m-%d")
        ws['A2'] = selected_date_obj.strftime("%B %Y")
        ws['A2'].font = Font(name='Arial', size=12)
        header_fill = PatternFill(start_color=self.colors['dark'].replace('#', ''),
                                 end_color=self.colors['dark'].replace('#', ''), fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        row = 4
        for col, header in enumerate(['Date', 'Total Purchase', 'Total Sales', 'Profit/Loss', 'Profit %'], 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20

    def update_summary_sheet(self, ws, wb):
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        row = 5
        total_purchase_sum = 0
        total_sales_sum = 0
        for sheet_name in wb.sheetnames:
            if sheet_name == "Monthly Summary":
                continue
            sheet = wb[sheet_name]
            total_purchase = 0
            total_sales = 0
            for r in sheet.iter_rows(min_row=1):
                if r[0].value == "Total Purchase:":
                    total_purchase = r[1].value if r[1].value else 0
                elif r[0].value == "Total Sales:":
                    total_sales = r[1].value if r[1].value else 0
            profit = total_sales - total_purchase
            profit_percent = (profit / total_purchase * 100) if total_purchase > 0 else 0
            ws.cell(row, 1, sheet_name).border = border
            ws.cell(row, 2, total_purchase).border = border
            ws.cell(row, 2).number_format = '"PKR "#,##0.00'
            ws.cell(row, 3, total_sales).border = border
            ws.cell(row, 3).number_format = '"PKR "#,##0.00'
            ws.cell(row, 4, profit).border = border
            ws.cell(row, 4).number_format = '"PKR "#,##0.00'
            ws.cell(row, 5, f"{profit_percent:.2f}%").border = border
            total_purchase_sum += total_purchase
            total_sales_sum += total_sales
            row += 1
        row += 1
        ws.cell(row, 1, "MONTHLY TOTAL").font = Font(bold=True, size=12)
        ws.cell(row, 2, total_purchase_sum).font = Font(bold=True, size=12)
        ws.cell(row, 2).number_format = '"PKR "#,##0.00'
        ws.cell(row, 3, total_sales_sum).font = Font(bold=True, size=12)
        ws.cell(row, 3).number_format = '"PKR "#,##0.00'
        ws.cell(row, 4, total_sales_sum - total_purchase_sum).font = Font(bold=True, size=12)
        ws.cell(row, 4).number_format = '"PKR "#,##0.00'
        monthly_profit_percent = ((total_sales_sum - total_purchase_sum) / total_purchase_sum * 100) if total_purchase_sum > 0 else 0
        ws.cell(row, 5, f"{monthly_profit_percent:.2f}%").font = Font(bold=True, size=12)

    def on_closing(self):
        try:
            self.all_purchases[self.selected_date] = self.purchases
            self.all_sales[self.selected_date] = self.sales
            self.save_all_purchases()
            self.save_all_sales()
            self.save_vegetables()
            self.save_invoice_counter()
            self.save_invoices()
            print("✅ All data saved before exit.")
        except Exception as e:
            print(f"⚠️ Final save failed: {e}")
            try:
                messagebox.showwarning("Save Warning", "Some data may not have been saved before closing.")
            except Exception:
                pass
        self.root.destroy()

if __name__ == "__main__":
    try:
        root = ctk.CTk()
        app = FruzyBusinessManager(root)
        root.mainloop()
    except Exception as e:
        try:
            messagebox.showerror("Fatal Error", f"Unhandled error: {e}")
        except Exception:
            print("Fatal Error:", e)
